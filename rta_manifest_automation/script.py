import openpyxl
from openpyxl.styles import Alignment, PatternFill, numbers
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
import re
import os
import time


ORANGE_FILL = PatternFill(start_color="FFFFD580",
                          end_color="FFFFD580", fill_type="solid")
GRAY_FILL = PatternFill(start_color='DDDDDD',
                        end_color='DDDDDD', fill_type='solid')
FILL_LIGHT_ORANGE = PatternFill(
    start_color="FFFFD580", end_color="FFFFD580", fill_type="solid")
FILL_LIGHT_PURPLE = PatternFill(
    start_color="FFE5CCFF", end_color="FFE5CCFF", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(
    start_color="FFCCEEFF", end_color="FFCCEEFF", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(
    start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")


def run_processing_pipeline(filepath):
    wb = process_step_1(filepath)
    process_step_2(wb)
    process_step_3(wb)
    process_step_4(wb)
    new_filename = generate_new_filename(filepath)
    wb.save(new_filename)
    print(f"Processing complete. File saved as: {new_filename}")


def process_step_1(filepath):
    wb = openpyxl.load_workbook(filepath)
    original_sheet = wb.sheetnames[0]
    step1_sheet = copy_sheet(wb, original_sheet, "Step 1")

    # if not verify_date_range(step1_sheet):
    #     raise ValueError("Date range in A2 is not for the current month")

    format_header(step1_sheet)
    highlight_rows(step1_sheet)
    return wb


def process_step_2(workbook):
    step2_sheet = copy_sheet(workbook, "Step 1", "Step 2")

    delete_above_header(step2_sheet)
    fix_column_headers_for_sheet_2(step2_sheet)
    format_header(step2_sheet, header_row=1)
    footer_row = get_footer_row(step2_sheet)

    col_totals = insert_mechanical_totals(step2_sheet, footer_row)
    validate_totals(step2_sheet, footer_row, col_totals)
    sum_formula_row = adjust_amount_total_with_deductions(
        step2_sheet, footer_row)
    finalize_adjusted_total_validation(step2_sheet, sum_formula_row)


def process_step_3(workbook):
    step3_sheet = copy_sheet(workbook, "Step 2", "Step 3")

    remove_empty_row_below_header(step3_sheet)
    format_header(step3_sheet, header_row=1)
    insert_time_column(step3_sheet)
    highlight_rows(step3_sheet, header_row=1)

    extract_time_from_item(step3_sheet)
    forward_fill_columns(step3_sheet)
    fill_customer_column_by_regid(step3_sheet)


def process_step_4(workbook):
    step4 = copy_sheet(workbook, "Step 3", "Step 4")

    remove_empty_columns(step4)
    remove_columns_by_header(step4, ["SubTotal", "Tax", "Total"])
    drop_rows_with_empty_item(step4)
    remove_footer_and_mech_rows(step4)
    clear_all_highlighting(step4)

    format_header(step4, header_row=1)
    highlight_rows(step4, header_row=1)
    distribute_items_to_sheets(step4, workbook)


# helpers
def generate_new_filename(filepath):
    base, ext = os.path.splitext(filepath)
    return f"{base} - Processed{ext}"


def copy_sheet(workbook, source_name, target_name):
    source = workbook[source_name]
    target = workbook.copy_worksheet(source)
    target.title = target_name
    return target


def highlight_row(sheet, row, max_col, fill):
    for col in range(1, max_col + 1):
        sheet.cell(row=row, column=col).fill = fill


def get_footer_row(sheet):
    for row in reversed(range(1, sheet.max_row + 1)):
        if any(sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)):
            return row
    raise ValueError("Footer row not found")


def format_header(sheet, header_row=7):
    if sheet[f"A{header_row}"].value != "RegID":
        raise ValueError(f"Expected 'RegID' in cell A{header_row}")

    sheet.freeze_panes = f"A{header_row + 1}"

    max_col = sheet.max_column
    max_row = sheet.max_row
    for col in range(1, max_col + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.alignment = Alignment(horizontal='left')
    last_col_letter = get_column_letter(max_col)
    sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{max_row}"


def highlight_rows(sheet, header_row=7):
    max_row = sheet.max_row
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        sheet.cell(row=header_row, column=col).fill = GRAY_FILL
    for row in reversed(range(1, max_row + 1)):
        if any(sheet.cell(row=row, column=col).value for col in range(1, max_col + 1)):
            for col in range(1, max_col + 1):
                sheet.cell(row=row, column=col).fill = GRAY_FILL
            break


def get_column_index_by_header(sheet, header_name, header_row=1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)
        if str(cell.value).strip().lower() == header_name.strip().lower():
            return col
    raise ValueError(f"Header '{header_name}' not found in row {header_row}")


def freeze_top_and_filter(sheet):
    sheet.freeze_panes = "A2"
    header_row = 1
    last_col = get_column_letter(sheet.max_column)
    last_row = sheet.max_row
    sheet.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


def sort_sheet_by_column(sheet, col_index, header_row, last_row):
    data = []
    for row in range(header_row + 1, last_row + 1):
        row_values = [sheet.cell(row=row, column=col).value for col in range(
            1, sheet.max_column + 1)]
        data.append((sheet.cell(row=row, column=col_index).value, row_values))

    data.sort(key=lambda x: (x[0] if x[0] is not None else float('inf')))

    for i, (_, row_values) in enumerate(data, start=header_row + 1):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=i, column=col_idx).value = value


def remove_empty_columns(sheet):
    for col in range(sheet.max_column, 0, -1):
        if all(sheet.cell(row=row, column=col).value in (None, "") for row in range(1, sheet.max_row+1)):
            sheet.delete_cols(col)


def remove_columns_by_header(sheet, headers):
    header_row = 1
    for header in headers:
        try:
            col = get_column_index_by_header(sheet, header, header_row)
            sheet.delete_cols(col)
        except ValueError:
            continue


def drop_rows_with_empty_item(sheet):
    item_col = get_column_index_by_header(sheet, "Item", 1)
    for row in range(sheet.max_row, 1, -1):
        if not sheet.cell(row=row, column=item_col).value:
            sheet.delete_rows(row)


def clear_all_highlighting(sheet):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = PatternFill()


def apply_filter_top(sheet):
    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A1:{last_col}{sheet.max_row}"


def remove_footer_and_mech_rows(sheet):
    footer_start = get_footer_row(sheet)
    sheet.delete_rows(footer_start, sheet.max_row - footer_start + 1)


def color_row(sheet, row, fill):
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=row, column=col).fill = fill


# step 1
def verify_date_range(sheet):
    date_text = sheet['A2'].value
    date_text = date_text.strip()
    match = re.match(
        r"(\d{2}/\d{2}/\d{2}) to (\d{2}/\d{2}/\d{2})", str(date_text))
    if not match:
        raise ValueError(f"Invalid date format in cell A2: {date_text}")
    start_str, _ = match.groups()
    start_date = datetime.strptime(start_str, "%m/%d/%y")
    now = datetime.now()
    return start_date.month == now.month and start_date.year == now.year


def delete_above_header(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=20):
        if row[0].value == "RegID":
            header_row = row[0].row
            for _ in range(header_row - 1):
                sheet.delete_rows(1)
            return
    raise ValueError("Header row not found")


# step 2
def insert_mechanical_totals(sheet, footer_row):
    sheet.insert_rows(footer_row, 2)
    mech_row = footer_row
    sheet[f"C{mech_row}"] = "Mechanical Totals"

    headers = ["Amount", "SubTotal", "Tax", "Total"]
    col_totals = {}

    for header in headers:
        col_index = get_column_index_by_header(sheet, header, header_row=1)
        col_letter = get_column_letter(col_index)

        start, end = 2, mech_row - 1
        formula = f"=SUM({col_letter}{start}:{col_letter}{end})"
        cell = sheet.cell(row=mech_row, column=col_index)
        cell.value = formula
        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # Calculate total using values directly
        total = 0
        for row in sheet.iter_rows(min_row=start, max_row=end, min_col=col_index, max_col=col_index):
            val = row[0].value
            if isinstance(val, (int, float)):
                total += val

        col_totals[col_letter] = total

    return col_totals


def validate_totals(sheet, footer_row, col_totals):
    headers = ["SubTotal", "Tax", "Total"]
    for header in headers:
        col_index = get_column_index_by_header(sheet, header, header_row=1)
        col_letter = get_column_letter(col_index)
        footer_val = sheet.cell(row=footer_row + 2, column=col_index).value
        calc_val = col_totals[col_letter]

        if footer_val is not None:
            try:
                clean = str(footer_val).replace(
                    "(", "-").replace(")", "").replace("$", "").replace(",", "")
                footer_float = float(clean)
                if round(calc_val, 2) != round(footer_float, 2):
                    raise ValueError(
                        f"Mismatch in column '{header}': {calc_val:.2f} != {footer_float:.2f}")
            except Exception:
                raise ValueError(
                    f"Footer value in '{header}' (row {footer_row + 2}) is invalid: {footer_val}")


def adjust_amount_total_with_deductions(sheet, mech_row):
    col_item = get_column_index_by_header(sheet, "Item")
    col_amount = get_column_index_by_header(sheet, "Amount")
    col_subtotal = get_column_index_by_header(sheet, "SubTotal")

    # Step 1: Adjust amount where SubTotal is negative
    for row in range(2, mech_row):
        sub_val = sheet.cell(row=row, column=col_subtotal).value
        if isinstance(sub_val, (int, float)) and sub_val < 0:
            next_row = row + 1
            amt_cell = sheet.cell(row=next_row, column=col_amount)
            amt_val = amt_cell.value
            if isinstance(amt_val, (int, float)):
                if amt_val > 0:
                    amt_cell.value = -amt_val
                highlight_row(sheet, row, sheet.max_column, ORANGE_FILL)

    # Step 2: Adjust amount where item contains keyword
    keywords = ["discount", "coupon", "petty"]
    for row in range(2, mech_row):
        item = sheet.cell(row=row, column=col_item).value
        if isinstance(item, str) and any(kw in item.lower() for kw in keywords):
            amt_cell = sheet.cell(row=row, column=col_amount)
            amt_val = amt_cell.value
            if isinstance(amt_val, (int, float)) and amt_val > 0:
                amt_cell.value = -amt_val
            if isinstance(amt_val, (int, float)):
                highlight_row(sheet, row, sheet.max_column, ORANGE_FILL)

    # Step 3: Insert sum formula for adjusted amount
    sum_formula_row = mech_row + 1
    amount_letter = get_column_letter(col_amount)
    sheet[f"{amount_letter}{sum_formula_row}"] = f"=SUM({amount_letter}2:{amount_letter}{mech_row - 1})"
    sheet[f"{amount_letter}{sum_formula_row}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    sheet[f"{get_column_letter(col_amount - 1)}{sum_formula_row}"] = "Difference"

    return sum_formula_row


def finalize_adjusted_total_validation(sheet, mech_row):
    data_start = 2
    data_end = mech_row - 1

    amount_col = get_column_index_by_header(sheet, "Amount")
    subtotal_col = get_column_index_by_header(sheet, "SubTotal")

    amount_sum = compute_column_sum_by_index(
        sheet, amount_col, data_start, data_end)
    subtotal_sum = compute_column_sum_by_index(
        sheet, subtotal_col, data_start, data_end)

    sheet.cell(row=mech_row + 1, column=amount_col).value = amount_sum
    sheet.cell(row=mech_row + 1,
               column=amount_col).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    if round(amount_sum, 2) != round(subtotal_sum, 2):
        raise ValueError(
            f"Adjusted Amount Total {amount_sum:.2f} != SubTotal {subtotal_sum:.2f}"
        )


def compute_column_sum_by_index(sheet, col_index, start_row, end_row):
    total = 0
    for row in range(start_row, end_row + 1):
        val = sheet.cell(row=row, column=col_index).value
        if isinstance(val, (int, float)):
            total += val
    return total


def fix_column_headers_for_sheet_2(sheet):
    if not sheet["C1"].value:
        sheet["C1"] = "Item"
    if not sheet["H1"].value:
        sheet["H1"] = "Amount"


# step 3
def insert_time_column(sheet):
    date_col = get_column_index_by_header(sheet, "Date/Time", header_row=1)
    time_col = date_col + 1

    sheet.insert_cols(time_col)
    sheet.cell(row=1, column=date_col).value = "Date"
    sheet.cell(row=1, column=time_col).value = "Time"


def extract_time_from_item(sheet):
    header_row = 1
    time_col = get_column_index_by_header(sheet, "Time", header_row)
    item_col = get_column_index_by_header(sheet, "Item", header_row)
    last_row = sheet.max_row

    time_pattern = re.compile(r"\b(\d{1,2}:\d{2}\s?(?:AM|PM|am|pm))\b")

    for row in range(2, last_row + 1):
        item_cell = sheet.cell(row=row, column=item_col)
        if item_cell.value and isinstance(item_cell.value, str):
            match = time_pattern.search(item_cell.value)
            if match:
                time_str = match.group(1)
                sheet.cell(row=row, column=time_col).value = time_str.strip()
                item_cell.value = item_cell.value.replace(time_str, "").strip()


def forward_fill_columns(sheet):
    header_row = 1
    mech_row = get_mechanical_totals_row(sheet)
    data_end_row = mech_row - 2

    columns_to_fill = ["RegID", "Date", "Time", "Tender"]
    for col_name in columns_to_fill:
        col_index = get_column_index_by_header(sheet, col_name, header_row)
        last_value = None
        for row in range(2, data_end_row + 1):
            cell = sheet.cell(row=row, column=col_index)
            if cell.value in (None, ""):
                if last_value is not None:
                    if col_name.lower() == "date":
                        # Ensure correct date format without time
                        if isinstance(last_value, datetime):
                            cell.value = last_value.date()
                        else:
                            cell.value = last_value
                        cell.number_format = 'm/d/yyyy'
                    else:
                        cell.value = last_value
            else:
                last_value = cell.value


def get_mechanical_totals_row(sheet):
    for row in range(1, sheet.max_row + 1):
        val = sheet.cell(row=row, column=get_column_index_by_header(
            sheet, "Item", 1)).value
        if isinstance(val, str) and "mechanical totals" in val.lower():
            return row
    raise ValueError("'Mechanical Totals' row not found.")


def remove_empty_row_below_header(sheet):
    row_idx = 2
    if all(sheet.cell(row=row_idx, column=col).value in (None, "") for col in range(1, sheet.max_column + 1)):
        sheet.delete_rows(row_idx)


def fill_customer_column_by_regid(sheet):
    header_row = 1
    mech_row = get_mechanical_totals_row(sheet)
    last_data_row = mech_row - 2

    regid_col = get_column_index_by_header(sheet, "RegID", header_row)
    customer_col = get_column_index_by_header(sheet, "Customer", header_row)

    # Step 1: Sort by RegID (ascending)
    sort_sheet_by_column(sheet, regid_col, header_row, last_data_row)

    # Step 2: Group-wise fill
    current_regid = None
    current_customer = None

    for row in range(2, last_data_row + 1):
        regid = sheet.cell(row=row, column=regid_col).value
        customer_cell = sheet.cell(row=row, column=customer_col)

        if regid != current_regid:
            current_regid = regid
            if customer_cell.value and str(customer_cell.value).strip():
                current_customer = customer_cell.value
            else:
                current_customer = "ZNoFirstName ZNoLastName"
                customer_cell.value = current_customer
        else:
            if customer_cell.value in (None, ""):
                customer_cell.value = current_customer


# step 4
def distribute_items_to_sheets(source, workbook):
    mapping = [
        ("DHL", "dhl", FILL_LIGHT_ORANGE, ["dhl drop off"]),
        ("USPS", "usps", FILL_LIGHT_PURPLE, ["void"]),
        ("FedEx", "fedex", FILL_LIGHT_BLUE,   ["void"]),
        ("UPS", "ups", FILL_LIGHT_GREEN,      ["void"])
    ]
    TAB_COLORS = {
        "DHL": "FFD580",   # Light Orange
        "USPS": "E5CCFF",  # Light Purple
        "FedEx": "CCEEFF",  # Light Blue
        "UPS": "CCFFCC"    # Light Green
    }
    item_col = get_column_index_by_header(source, "Item", 1)

    for sheet_name, keyword, fill, excludes in mapping:
        target = workbook.create_sheet(sheet_name)
        target.sheet_properties.tabColor = TAB_COLORS[sheet_name]
        copy_headers(source, target)
        format_header(target, header_row=1)
        freeze_top_and_filter(target)
        highlight_rows(target, header_row=1)

        tgt_row = 2
        for row in range(2, source.max_row + 1):
            val = str(source.cell(row=row, column=item_col).value or "")
            if keyword.lower() in val.lower() and not any(ex in val.lower() for ex in excludes):
                color_row(source, row, fill)
                copy_row_with_fill(source, target, row, tgt_row, fill)
                tgt_row += 1


def copy_headers(src, tgt):
    for col in range(1, src.max_column + 1):
        tgt.cell(row=1, column=col).value = src.cell(row=1, column=col).value


def copy_row_with_fill(src, tgt, src_row, tgt_row, fill):
    for col in range(1, src.max_column + 1):
        src_cell = src.cell(row=src_row, column=col)
        tgt_cell = tgt.cell(row=tgt_row, column=col)

        tgt_cell.value = src_cell.value
        tgt_cell.number_format = src_cell.number_format  # 🔥 key fix
        tgt_cell.fill = fill


start = time.time()
run_processing_pipeline("files/Processing/25 Feb.xlsx")
end = time.time()
print("Time : ", end - start)
