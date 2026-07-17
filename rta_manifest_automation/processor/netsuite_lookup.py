"""
NetSuite lookup module.
Loads the Item Master and Customer files, builds fast lookup dicts,
and provides helpers to resolve Item and Customer NetSuite IDs.
"""
import os
import csv
import openpyxl


class NetSuiteLookup:
    """
    Holds fast lookups for Item -> (NS ID, NS Name) and Customer -> NS ID.
    Use .lookup_item(item_text) and .lookup_customer(customer_text).
    Returns (None, None) or None when no match.
    """

    def __init__(self, item_master_path=None, customer_path=None):
        self.item_map = {}                        # normalized item -> (ns_id, ns_name)
        self.customer_map = {}                    # normalized rta_customer -> ns_id (first-seen)
        self.customers_with_multiple_ids = set()  # normalized rta_customer -> has 2+ NS IDs

        if item_master_path and os.path.exists(item_master_path):
            self._load_item_master(item_master_path)
            print(f"  [NSLookup] loaded {len(self.item_map)} items from master file")

        if customer_path and os.path.exists(customer_path):
            self._load_customer_file(customer_path)
            print(f"  [NSLookup] loaded {len(self.customer_map)} customers")

    @staticmethod
    def _normalize(text):
        if text is None:
            return ""
        # Lowercase + collapse all whitespace to single spaces + strip
        s = str(text).strip().lower()
        s = " ".join(s.split())
        return s

    def _load_item_master(self, path):
        """
        Load ONLY the first sheet. Expects columns:
        'Key Word', 'NetSuite Internal ID', 'NetSuite Name'
        (Column positions detected by header name — order doesn't matter)
        """
        if path.lower().endswith(".csv"):
            self._load_item_master_csv(path)
        else:
            self._load_item_master_xlsx(path)

    def _load_item_master_xlsx(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]  # first sheet only

        # Find column positions from header row
        header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
        col_key = col_id = col_name = None
        for i, h in enumerate(header):
            if h == "key word":
                col_key = i
            elif h == "netsuite internal id":
                col_id = i
            elif h == "netsuite name":
                col_name = i

        if col_key is None or col_id is None or col_name is None:
            print(f"  [NSLookup] WARNING: Item Master missing required columns. Headers found: {header}")
            wb.close()
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(col_key, col_id, col_name):
                continue
            key = row[col_key]
            ns_id = row[col_id]
            ns_name = row[col_name]
            if key is None:
                continue
            norm = self._normalize(key)
            if norm:
                # First occurrence wins; don't overwrite
                if norm not in self.item_map:
                    self.item_map[norm] = (ns_id, ns_name)
        wb.close()

    def _load_item_master_csv(self, path):
        with open(path, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = row.get("Key Word") or row.get("key word")
                ns_id = row.get("NetSuite Internal ID") or row.get("netsuite internal id")
                ns_name = row.get("NetSuite Name") or row.get("netsuite name")
                if not key:
                    continue
                norm = self._normalize(key)
                if norm and norm not in self.item_map:
                    self.item_map[norm] = (ns_id, ns_name)

    def _load_customer_file(self, path):
        """
        Expects a column named 'updated_rta' (match target) and 'primary_customer_ns_id'.
        """
        if path.lower().endswith(".csv"):
            self._load_customer_csv(path)
        else:
            self._load_customer_xlsx(path)

    def _load_customer_csv(self, path):
        with open(path, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rta_name = row.get("updated_rta")
                ns_id = row.get("primary_customer_ns_id")
                if not rta_name:
                    continue
                norm = self._normalize(rta_name)
                if not norm:
                    continue
                if norm in self.customer_map:
                    # Already seen this RTA name → it has multiple NS IDs
                    # Only flag if the ID is actually different (not just a duplicate blank row)
                    if ns_id and ns_id != self.customer_map[norm]:
                        self.customers_with_multiple_ids.add(norm)
                else:
                    self.customer_map[norm] = ns_id

    def _load_customer_xlsx(self, path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]

        header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
        col_rta = col_id = None
        for i, h in enumerate(header):
            if h == "updated_rta":
                col_rta = i
            elif h == "primary_customer_ns_id":
                col_id = i

        if col_rta is None or col_id is None:
            print(f"  [NSLookup] WARNING: Customer file missing required columns. Headers: {header}")
            wb.close()
            return

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(col_rta, col_id):
                continue
            rta_name = row[col_rta]
            ns_id = row[col_id]
            if not rta_name:
                continue
            norm = self._normalize(rta_name)
            if not norm:
                continue
            if norm in self.customer_map:
                if ns_id and ns_id != self.customer_map[norm]:
                    self.customers_with_multiple_ids.add(norm)
            else:
                self.customer_map[norm] = ns_id
        wb.close()

    
    def has_multiple_ids(self, customer_text):
        """Returns True if this customer has 2+ NS IDs in the source file."""
        if not customer_text:
            return False
        norm = self._normalize(customer_text)
        return norm in self.customers_with_multiple_ids

    def lookup_item(self, item_text):
        """Returns (ns_id, ns_name) or (None, None) if not found."""
        if not item_text:
            return (None, None)
        norm = self._normalize(item_text)
        return self.item_map.get(norm, (None, None))

    def lookup_customer(self, customer_text):
        """Returns ns_id or None."""
        if not customer_text:
            return None
        norm = self._normalize(customer_text)
        return self.customer_map.get(norm)