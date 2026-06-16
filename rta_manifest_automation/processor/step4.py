import time
from copy import copy
from openpyxl.styles import PatternFill, Font
from .utils import *
from datetime import date

FILL_LIGHT_ORANGE = PatternFill(start_color="FFFFD580", end_color="FFFFD580", fill_type="solid")
FILL_LIGHT_PURPLE = PatternFill(start_color="FFE5CCFF", end_color="FFE5CCFF", fill_type="solid")
FILL_LIGHT_BLUE   = PatternFill(start_color="FFCCEEFF", end_color="FFCCEEFF", fill_type="solid")
FILL_LIGHT_GREEN  = PatternFill(start_color="FFCCFFCC", end_color="FFCCFFCC", fill_type="solid")


def process_step_4(workbook):
    step4 = copy_sheet(workbook, "Step 3", "Step 4")
    remove_empty_columns(step4)
    remove_columns_by_header(step4, ["SubTotal", "Tax", "Total", "User"])
    drop_rows_with_empty_item(step4)
    remove_footer_and_mech_rows(step4)
    remove_mechanical_totals_row(step4)   # ← NEW: explicitly remove Mechanical Totals & Difference rows
    clear_all_highlighting(step4)
    format_header(step4, header_row=1)
    highlight_header_row(step4, header_row=1)
    autofit_columns(step4)
    add_uid_column(step4)
    distribute_items_to_sheets(step4, workbook)


def remove_mechanical_totals_row(sheet):
    """
    Explicitly find and remove any row containing 'Mechanical Totals' or 'Difference'.
    These rows come from Step 2 and should not appear in Step 4 onwards.
    """
    rows_to_delete = []
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row_cells:
            val = cell.value
            if isinstance(val, str):
                vl = val.strip().lower()
                if "mechanical total" in vl or vl == "difference":
                    rows_to_delete.append(row_cells[0].row)
                    break

    # Delete bottom-up to avoid index shifting
    for row_num in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_num, 1)


def add_uid_column(sheet):
    sheet.insert_cols(1)
    sheet.cell(row=1, column=1).value = "UID"

    date_col = get_column_index_by_header(sheet, "Date", 1)
    last_date = None
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                                      min_col=date_col, max_col=date_col):
        v = row_cells[0].value
        if v:
            last_date = v

    if last_date is None:
        last_date = date.today()
    if hasattr(last_date, 'strftime'):
        prefix = last_date.strftime("%y%m%d")
    else:
        from datetime import datetime
        prefix = datetime.strptime(str(last_date), "%Y-%m-%d").strftime("%y%m%d")

    max_col = sheet.max_column
    counter = 1
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row,
                                      min_col=2, max_col=max_col):
        if any(c.value for c in row_cells):
            sheet.cell(row=row_cells[0].row, column=1).value = int(f"{prefix}{counter:04d}")
            counter += 1


def distribute_items_to_sheets(source, workbook):
    mapping = [
        ("DHL",   "dhl",   FILL_LIGHT_ORANGE, ["dhl drop off"]),
        ("USPS",  "usps",  FILL_LIGHT_PURPLE, ["void"]),
        ("FedEx", "fedex", FILL_LIGHT_BLUE,   ["void"]),
        ("UPS",   "ups",   FILL_LIGHT_GREEN,  ["void"])
    ]
    TAB_COLORS = {"DHL": "FFD580", "USPS": "E5CCFF", "FedEx": "CCEEFF", "UPS": "CCFFCC"}

    item_col     = get_column_index_by_header(source, "Item", 1)
    regid_col    = get_column_index_by_header(source, "RegID", 1)
    customer_col = get_column_index_by_header(source, "Customer", 1)
    amount_col   = get_column_index_by_header(source, "Amount", 1)
    max_col = source.max_column
    max_row = source.max_row

    # Load source into memory
    src_data = []
    src_formats = []
    for row in source.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        src_data.append([c.value for c in row])
        src_formats.append([c.number_format for c in row])

    # RegID -> indices map for fast declared-value fallback
    regid_to_indices = {}
    for idx in range(1, len(src_data)):
        rid = src_data[idx][regid_col - 1]
        if rid is not None:
            regid_to_indices.setdefault(rid, []).append(idx)

    # Create service sheets with headers
    service_sheets = {}
    for sheet_name, keyword, fill, excludes in mapping:
        target = workbook.create_sheet(sheet_name)
        target.sheet_properties.tabColor = TAB_COLORS[sheet_name]
        for col_idx, val in enumerate(src_data[0], start=1):
            target.cell(row=1, column=col_idx).value = val
        service_sheets[sheet_name] = {"sheet": target, "fill": fill, "rows_buffer": []}

    # Classification pass
    i = 1
    while i < len(src_data):
        row_data = src_data[i]
        row_fmts = src_formats[i]
        item_val = str(row_data[item_col - 1] or "")
        item_lower = item_val.lower()
        regid = row_data[regid_col - 1]
        customer = str(row_data[customer_col - 1] or "")
        next_idx = i + 1

        # Declared value pairing
        if "declared value" in item_lower:
            handled = False
            if next_idx < len(src_data):
                next_data = src_data[next_idx]
                next_item = str(next_data[item_col - 1] or "").lower()
                next_regid = next_data[regid_col - 1]
                if regid == next_regid:
                    for sheet_name, keyword, fill, excludes in mapping:
                        if keyword in next_item and not any(ex in next_item for ex in excludes):
                            service_sheets[sheet_name]["rows_buffer"].append((row_data, row_fmts))
                            handled = True
                            break
            if not handled and regid in regid_to_indices:
                for search_idx in regid_to_indices[regid]:
                    if search_idx == i:
                        continue
                    search_item = str(src_data[search_idx][item_col - 1] or "").lower()
                    for sheet_name, keyword, fill, excludes in mapping:
                        if keyword in search_item and not any(ex in search_item for ex in excludes):
                            service_sheets[sheet_name]["rows_buffer"].append((row_data, row_fmts))
                            handled = True
                            break
                    if handled:
                        break
            i += 1
            continue

        # Service row check
        matched_service = None
        for sheet_name, keyword, fill, excludes in mapping:
            if keyword in item_lower and not any(ex in item_lower for ex in excludes):
                matched_service = (sheet_name, keyword, fill, excludes)
                break

        if matched_service is None:
            i += 1
            continue

        sheet_name, keyword, fill, excludes = matched_service
        service_sheets[sheet_name]["rows_buffer"].append((row_data, row_fmts))

        # Existing discount/coupon
       # Existing discount/coupon — keep grabbing consecutive ones tied to same RegID
        has_existing_discount = False
        scan_idx = next_idx
        while scan_idx < len(src_data):
            scan_data = src_data[scan_idx]
            scan_fmts = src_formats[scan_idx]
            scan_item = str(scan_data[item_col - 1] or "").lower()
            scan_regid = scan_data[regid_col - 1]

            # Must be same RegID
            if scan_regid != regid:
                break

            # Must be discount/coupon (and not void)
            if not (("discount" in scan_item or "coupon" in scan_item) and "void" not in scan_item):
                break

            # Append and continue scanning
            service_sheets[sheet_name]["rows_buffer"].append((scan_data, scan_fmts))
            has_existing_discount = True
            scan_idx += 1

        # Empire 50% discount
        if customer.strip().lower() == "empire merchants chelsea" and not has_existing_discount:
            discount_data = list(row_data)
            discount_fmts = list(row_fmts)
            discount_data[item_col - 1] = "50% discount"
            orig_amount = row_data[amount_col - 1]
            if isinstance(orig_amount, (int, float)):
                discount_data[amount_col - 1] = -abs(orig_amount) / 2
            service_sheets[sheet_name]["rows_buffer"].append((discount_data, discount_fmts))

        i += 1

    # Write buffers to service sheets
    for sheet_name, info in service_sheets.items():
        target = info["sheet"]
        fill = info["fill"]
        write_row = 2
        for row_data, row_fmts in info["rows_buffer"]:
            for col_idx, (val, fmt) in enumerate(zip(row_data, row_fmts), start=1):
                cell = target.cell(row=write_row, column=col_idx)
                cell.value = val
                cell.fill = fill
                if fmt:
                    cell.number_format = fmt
            write_row += 1
        format_header(target, header_row=1)
        freeze_top_and_filter(target)
        highlight_header_row(target, header_row=1)
        autofit_columns(target)

    build_3pl_sheet(workbook, service_sheets)
    build_account_sheets(source, workbook, mapping)


def build_3pl_sheet(workbook, service_sheets):
    SERVICE_ORDER = ["DHL", "USPS", "FedEx", "UPS"]
    EXCLUDED_TENDER = "account"
    EXCLUDED_ITEM_KEYWORDS = ["void"]

    if "3PL" in workbook.sheetnames:
        del workbook["3PL"]
    sheet_3pl = workbook.create_sheet("3PL")

    first_sheet = None
    for sname in SERVICE_ORDER:
        if sname in workbook.sheetnames:
            first_sheet = workbook[sname]
            break
    if first_sheet is None:
        return

    header_values = []
    tender_col = item_col_3pl = uid_col = None
    for col in range(1, first_sheet.max_column + 1):
        val = first_sheet.cell(row=1, column=col).value
        header_values.append(val)
        hv = str(val or "").strip().lower()
        if hv == "tender":
            tender_col = col
        elif hv == "item":
            item_col_3pl = col
        elif hv == "uid":
            uid_col = col

    for col_idx, val in enumerate(header_values, start=1):
        sheet_3pl.cell(row=1, column=col_idx).value = val

    write_row = 2
    for sname in SERVICE_ORDER:
        if sname not in service_sheets:
            continue
        info = service_sheets[sname]
        fill = info["fill"]
        for row_data, row_fmts in info["rows_buffer"]:
            if tender_col:
                tv = str(row_data[tender_col - 1] or "").strip().lower()
                if tv == EXCLUDED_TENDER:
                    continue
            if item_col_3pl:
                iv = str(row_data[item_col_3pl - 1] or "").strip().lower()
                if any(kw in iv for kw in EXCLUDED_ITEM_KEYWORDS):
                    continue
            if uid_col:
                if not row_data[uid_col - 1]:
                    continue
            for col_idx, (val, fmt) in enumerate(zip(row_data, row_fmts), start=1):
                cell = sheet_3pl.cell(row=write_row, column=col_idx)
                cell.value = val
                cell.fill = fill
                if fmt:
                    cell.number_format = fmt
            write_row += 1

    format_header(sheet_3pl, header_row=1)
    freeze_top_and_filter(sheet_3pl)
    highlight_header_row(sheet_3pl, header_row=1)
    autofit_columns(sheet_3pl)


def build_account_sheets(source, workbook, mapping):
    ACCOUNT_CUSTOMERS = [
        ("E-Scribers", ["e-scriber", "escriber"]),
        ("Empire",     ["empire"]),
        ("Feshaire",   ["feshaire", "fashaire"]),
    ]
    EXCLUDED_ITEM_KEYWORDS = ["void"]

    item_col     = get_column_index_by_header(source, "Item", 1)
    tender_col   = get_column_index_by_header(source, "Tender", 1)
    customer_col = get_column_index_by_header(source, "Customer", 1)
    amount_col   = get_column_index_by_header(source, "Amount", 1)
    regid_col    = get_column_index_by_header(source, "RegID", 1)
    max_col      = source.max_column

    src_data = []
    src_formats = []
    for row in source.iter_rows(min_row=1, max_row=source.max_row, max_col=max_col):
        src_data.append([c.value for c in row])
        src_formats.append([c.number_format for c in row])

    def get_service_fill(item_val):
        item_lower = item_val.lower()
        for sheet_name, keyword, fill, excludes in mapping:
            if keyword.lower() in item_lower and not any(ex in item_lower for ex in excludes):
                return fill
        return FILL_LIGHT_GREEN

    for sheet_name, customer_keywords in ACCOUNT_CUSTOMERS:
        is_empire = (sheet_name == "Empire")

        matching_indices = []
        for idx in range(1, len(src_data)):
            row_data = src_data[idx]
            tender_val   = str(row_data[tender_col - 1] or "").strip().lower()
            customer_val = str(row_data[customer_col - 1] or "").lower()
            item_val     = str(row_data[item_col - 1] or "").lower()

            if tender_val != "account":
                continue
            if not any(kw in customer_val for kw in customer_keywords):
                continue
            if any(kw in item_val for kw in EXCLUDED_ITEM_KEYWORDS):
                continue
            matching_indices.append(idx)

        if not matching_indices:
            continue

        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)

        for col_idx, val in enumerate(src_data[0], start=1):
            ws.cell(row=1, column=col_idx).value = val

        write_row = 2
        i = 0
        while i < len(matching_indices):
            src_idx = matching_indices[i]
            row_data = src_data[src_idx]
            row_fmts = src_formats[src_idx]
            item_val   = str(row_data[item_col - 1] or "")
            item_lower = item_val.lower()
            amount_val = row_data[amount_col - 1]
            row_fill   = get_service_fill(item_val)

            for col_idx, (val, fmt) in enumerate(zip(row_data, row_fmts), start=1):
                tgt_cell = ws.cell(row=write_row, column=col_idx)
                tgt_cell.value = val
                if fmt:
                    tgt_cell.number_format = fmt
                tgt_cell.fill = row_fill
            write_row += 1

            if is_empire:
                is_service_row = (
                    "discount" not in item_lower and
                    "coupon" not in item_lower and
                    "void" not in item_lower
                )
                if is_service_row:
                    regid = row_data[regid_col - 1]
                    next_is_discount = False
                    if i + 1 < len(matching_indices):
                        next_src_idx = matching_indices[i + 1]
                        next_data = src_data[next_src_idx]
                        next_item = str(next_data[item_col - 1] or "").lower()
                        next_regid = next_data[regid_col - 1]
                        if next_regid == regid and ("discount" in next_item or "coupon" in next_item):
                            next_is_discount = True

                    if not next_is_discount:
                        for col_idx, (val, fmt) in enumerate(zip(row_data, row_fmts), start=1):
                            tgt_cell = ws.cell(row=write_row, column=col_idx)
                            tgt_cell.value = val
                            if fmt:
                                tgt_cell.number_format = fmt
                            tgt_cell.fill = row_fill
                        ws.cell(row=write_row, column=1).value = None
                        ws.cell(row=write_row, column=item_col).value = "50% discount"
                        if isinstance(amount_val, (int, float)):
                            ws.cell(row=write_row, column=amount_col).value = -abs(amount_val) / 2
                        write_row += 1

            i += 1

        # Net total
        total_row = write_row + 1
        # Net total — sum directly from written rows (includes auto-generated 50% discounts)
        net_total = 0.0
        for r in range(2, write_row):
            val = ws.cell(row=r, column=amount_col).value
            if isinstance(val, (int, float)):
                net_total += val

        total_label_col = amount_col - 1 if amount_col > 1 else amount_col
        ws.cell(row=total_row, column=total_label_col).value = "Total:"
        ws.cell(row=total_row, column=total_label_col).font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=amount_col)
        total_cell.value = round(net_total, 2)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '$#,##0.00'

        format_header(ws, header_row=1)
        freeze_top_and_filter(ws)
        highlight_header_row(ws, header_row=1)
        autofit_columns(ws)