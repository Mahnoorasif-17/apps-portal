import openpyxl
from datetime import datetime
import re
from .utils import *


def process_step_1(filepath):
    wb = openpyxl.load_workbook(filepath)
    original_sheet = wb.sheetnames[0]
    step1_sheet = copy_sheet(wb, original_sheet, "Step 1")

    # if not verify_date_range(step1_sheet):
    #     raise ValidationError("Date range in A2 is not for the current month")

    format_header(step1_sheet)
    highlight_rows(step1_sheet)

    # --- Fix rows where SubTotal/Tax/Total are shifted 1 column right ---
    fix_shifted_money_columns(step1_sheet)

    # --- Delete trailing empty columns AFTER shifting is done ---
    delete_trailing_empty_columns(step1_sheet)

    # --- Reset filter range to match the trimmed sheet ---
    reset_filter_range(step1_sheet)

    return wb


def fix_shifted_money_columns(sheet):
    """
    Detects rows where SubTotal is empty but Tax, Total, and the column after Total
    all have money values. Shifts those values back into SubTotal/Tax/Total,
    preserving number formats.
    """
    # Find header row (where "RegID" is)
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=20):
        if row[0].value == "RegID":
            header_row = row[0].row
            break
    if header_row is None:
        return

    try:
        subtotal_col = get_column_index_by_header(sheet, "SubTotal", header_row)
        tax_col      = get_column_index_by_header(sheet, "Tax", header_row)
        total_col    = get_column_index_by_header(sheet, "Total", header_row)
    except Exception:
        return

    max_row = sheet.max_row
    max_col = sheet.max_column
    col_after_total = total_col + 1

    # Get the standard currency format from any existing properly-formatted money cell
    # Default fallback if we can't find one
    CURRENCY_FORMAT = '"$"#,##0.00'

    fixed_count = 0

    for row_idx in range(header_row + 1, max_row + 1):
        if col_after_total > max_col:
            break

        subtotal_cell    = sheet.cell(row=row_idx, column=subtotal_col)
        tax_cell         = sheet.cell(row=row_idx, column=tax_col)
        total_cell       = sheet.cell(row=row_idx, column=total_col)
        after_total_cell = sheet.cell(row=row_idx, column=col_after_total)

        subtotal_val    = subtotal_cell.value
        tax_val         = tax_cell.value
        total_val       = total_cell.value
        after_total_val = after_total_cell.value

        # Detect shift-by-1
        if (subtotal_val in (None, "") and
            isinstance(tax_val, (int, float)) and
            isinstance(total_val, (int, float)) and
            isinstance(after_total_val, (int, float))):

            # Capture the formats BEFORE shifting (so we move them too)
            tax_fmt         = tax_cell.number_format
            total_fmt       = total_cell.number_format
            after_total_fmt = after_total_cell.number_format

            # Shift values left by 1
            subtotal_cell.value = tax_val
            tax_cell.value      = total_val
            total_cell.value    = after_total_val
            after_total_cell.value = None

            # Shift formats too (so $ sign stays)
            subtotal_cell.number_format = tax_fmt if tax_fmt and tax_fmt != "General" else CURRENCY_FORMAT
            tax_cell.number_format      = total_fmt if total_fmt and total_fmt != "General" else CURRENCY_FORMAT
            total_cell.number_format    = after_total_fmt if after_total_fmt and after_total_fmt != "General" else CURRENCY_FORMAT

            fixed_count += 1

    if fixed_count > 0:
        print(f"  [Step1] shifted {fixed_count} rows back into SubTotal/Tax/Total columns")
def delete_trailing_empty_columns(sheet):
    """
    Deletes any trailing columns that are completely empty (no header text,
    no data in any row). Stops at the last column that has any content.

    Note: only deletes columns AFTER the last column with a header label.
    Empty columns between data columns (like the unnamed col between Customer
    and SubTotal in your source) are NOT touched.
    """
    # Find the last column index that has a header value
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=20):
        if row[0].value == "RegID":
            header_row = row[0].row
            break
    if header_row is None:
        return

    # Find last column with a header label
    last_labeled_col = 0
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val not in (None, ""):
            last_labeled_col = col

    if last_labeled_col == 0:
        return

    # Delete every column AFTER the last labeled one, going right-to-left
    deleted = 0
    for col in range(sheet.max_column, last_labeled_col, -1):
        # Safety: only delete if entire column is truly empty
        has_data = False
        for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row,
                                          min_col=col, max_col=col):
            if row_cells[0].value not in (None, ""):
                has_data = True
                break
        if not has_data:
            sheet.delete_cols(col)
            deleted += 1

    if deleted > 0:
        print(f"  [Step1] deleted {deleted} trailing empty columns")


def reset_filter_range(sheet):
    """
    After deleting trailing columns, re-set the auto_filter range
    so the chevron doesn't extend past the actual data.
    """
    from openpyxl.utils import get_column_letter

    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=20):
        if row[0].value == "RegID":
            header_row = row[0].row
            break
    if header_row is None:
        return

    last_col_letter = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{sheet.max_row}"


def verify_date_range(sheet):
    date_text = sheet['A2'].value
    date_text = date_text.strip()
    match = re.match(
        r"(\d{2}/\d{2}/\d{2}) to (\d{2}/\d{2}/\d{2})", str(date_text))
    if not match:
        raise ValidationError(f"Invalid date format in cell A2: {date_text}")
    start_str, _ = match.groups()
    start_date = datetime.strptime(start_str, "%m/%d/%y")
    now = datetime.now()
    return start_date.month == now.month and start_date.year == now.year


def delete_above_header(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=20):
        if row[0].value == "RegID":
            header_row = row[0].row
            for _ in range(header_row - 1):
                sheet.delete_rows(1)
            return
    raise ValidationError("Header row not found")