import re
import time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


DUP_YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def clean_text(val):
    if val is None:
        return ""
    return re.sub(r"[^a-zA-Z0-9 ]", " ", str(val)).lower().strip()


def parse_amount(val):
    if val is None or val == "":
        return 0.0
    val = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(val)
    except:
        return 0.0


def process_step_6(workbook):
    t0 = time.time()

    step5 = workbook["Step 5"]

    if "Step 6" in workbook.sheetnames:
        del workbook["Step 6"]

    step6 = workbook.create_sheet("Step 6")

    # New headers include 3 NS columns at the end
    headers = [
        "UID", "RegID", "Date", "Time", "Item",
        "Tender", "Customer", "Amount", "Tax", "Total Amount", "Taxable",
        "Item NetSuite ID", "Item NetSuite Name", "Customer NetSuite ID"
    ]
    step6.append(headers)

    item_col   = 5
    amount_col = 8

    TAX_RATE = 0.08875

    EXCLUDE_KEYWORDS = [
        "coupon", "discount", "void", "term",
        "late fee", "mailbox", "setup fee", "renew",
        "ups", "usps", "fedex", "dhl",
        "declared value",
    ]

    INCLUDE_OVERRIDES = ["dhl drop off"]

    TAXABLE_KEYWORDS = [
        "copies", "misc  taxable", "fax", "lamination",
        "passport", "postcard",
        "printing", "scan", "office rental"
    ]

    COUPON_DISCOUNT_KEYWORDS = ["coupon", "discount"]

    def is_no_fill_from_cell(cell):
        fill = cell.fill
        if fill is None:
            return True
        if fill.fill_type is None:
            return True
        if getattr(fill, "patternType", None) is None:
            return True
        return False

    # --- Find NS column positions in Step 5 by header name ---
    ns_item_id_src = ns_item_name_src = ns_cust_id_src = None
    for col in range(1, step5.max_column + 1):
        h = str(step5.cell(row=1, column=col).value or "").strip().lower()
        if h == "item netsuite id":
            ns_item_id_src = col
        elif h == "item netsuite name":
            ns_item_name_src = col
        elif h == "customer netsuite id":
            ns_cust_id_src = col

    # --- Load Step 5 into memory (values + fill info + NS cust ID cell fills) ---
    t = time.time()
    rows_values = []
    rows_no_fill = []
    rows_ns_cust_yellow = []  # True if the source NS Customer ID cell was yellow (duplicate)
    max_col = step5.max_column

    for row_cells in step5.iter_rows(min_row=1, max_row=step5.max_row, max_col=max_col):
        rows_values.append([c.value for c in row_cells])
        if len(row_cells) >= item_col:
            rows_no_fill.append(is_no_fill_from_cell(row_cells[item_col - 1]))
        else:
            rows_no_fill.append(True)

        # Check if the Customer NetSuite ID cell in this row was yellow
        is_yellow = False
        if ns_cust_id_src is not None and len(row_cells) >= ns_cust_id_src:
            cell = row_cells[ns_cust_id_src - 1]
            if cell.fill and cell.fill.fill_type == "solid":
                color = cell.fill.start_color.rgb if cell.fill.start_color else None
                if color and "FFFF00" in str(color).upper():
                    is_yellow = True
        rows_ns_cust_yellow.append(is_yellow)

    print(f"  [Step6] loaded {len(rows_values)} rows in {time.time()-t:.2f}s")

    processed_rows = []
    total_amount = 0.0
    total_tax    = 0.0
    total_total  = 0.0

    last_item_was_retail = False

    t = time.time()
    for r_idx in range(1, len(rows_values)):
        row_vals = rows_values[r_idx]

        item_raw   = row_vals[item_col - 1] if len(row_vals) >= item_col else None
        item_clean = clean_text(item_raw)
        amount     = parse_amount(row_vals[amount_col - 1] if len(row_vals) >= amount_col else None)

        if item_clean == "" and amount == 0:
            continue

        if not rows_no_fill[r_idx]:
            continue

        is_coupon_discount = any(k in item_clean for k in COUPON_DISCOUNT_KEYWORDS)

        # Extract NS values for this source row
        ns_item_id   = row_vals[ns_item_id_src - 1]   if ns_item_id_src   is not None and len(row_vals) >= ns_item_id_src   else None
        ns_item_name = row_vals[ns_item_name_src - 1] if ns_item_name_src is not None and len(row_vals) >= ns_item_name_src else None
        ns_cust_id   = row_vals[ns_cust_id_src - 1]   if ns_cust_id_src   is not None and len(row_vals) >= ns_cust_id_src   else None
        cust_is_yellow = rows_ns_cust_yellow[r_idx]

        if is_coupon_discount:
            if last_item_was_retail:
                row_data = []
                for c in range(1, 11):
                    if c <= len(row_vals):
                        row_data.append(row_vals[c - 1])
                    else:
                        row_data.append(None)

                row_data[7] = amount
                row_data[8] = 0.0
                row_data[9] = amount

                # Append taxable_flag + NS values
                full_row = row_data + ["n", ns_item_id, ns_item_name, ns_cust_id]
                processed_rows.append((item_clean, full_row, cust_is_yellow))

                total_amount += amount
                total_total  += amount
            continue

        is_override = any(ov in item_clean for ov in INCLUDE_OVERRIDES)

        if not is_override and any(k in item_clean for k in EXCLUDE_KEYWORDS):
            last_item_was_retail = False
            continue

        if "mechanical total" in item_clean:
            last_item_was_retail = False
            continue

        is_taxable   = any(k in item_clean for k in TAXABLE_KEYWORDS)
        tax          = round(amount * TAX_RATE, 2) if is_taxable else 0.0
        taxable_flag = "y" if is_taxable else "n"
        total        = round(amount + tax, 2)

        row_data = []
        for c in range(1, 11):
            if c <= len(row_vals):
                row_data.append(row_vals[c - 1])
            else:
                row_data.append(None)

        row_data[7] = amount
        row_data[8] = tax
        row_data[9] = total

        full_row = row_data + [taxable_flag, ns_item_id, ns_item_name, ns_cust_id]
        processed_rows.append((item_clean, full_row, cust_is_yellow))

        total_amount += amount
        total_tax    += tax
        total_total  += total

        last_item_was_retail = True

    print(f"  [Step6] processed {len(processed_rows)} rows in {time.time()-t:.2f}s")

    processed_rows.sort(key=lambda x: x[0])

    # Column indices in Step 6 output
    COL_AMOUNT_OUT = 8
    COL_TAX_OUT    = 9
    COL_TOTAL_OUT  = 10
    COL_TAXABLE    = 11
    COL_NS_CUST_ID_OUT = 14  # Customer NetSuite ID

    for _, full_row, cust_is_yellow in processed_rows:
        step6.append(full_row)
        # Highlight NS Customer ID cell yellow if the source was flagged
        if cust_is_yellow:
            step6.cell(row=step6.max_row, column=COL_NS_CUST_ID_OUT).fill = DUP_YELLOW

    for row in range(2, step6.max_row + 1):
        for col in [COL_AMOUNT_OUT, COL_TAX_OUT, COL_TOTAL_OUT]:
            cell = step6.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'

    step6.append([])
    step6.append([
        "", "", "", "", "", "", "TOTALS",
        total_amount,
        total_tax,
        total_total,
        "", "", "", ""
    ])

    last_row = step6.max_row
    for col in [COL_AMOUNT_OUT, COL_TAX_OUT, COL_TOTAL_OUT]:
        step6.cell(row=last_row, column=col).number_format = '$#,##0.00'
        step6.cell(row=last_row, column=col).font          = Font(bold=True)
    step6.cell(row=last_row, column=7).font = Font(bold=True)

    for col in range(1, 15):
        step6.column_dimensions[get_column_letter(col)].width = 20

    step6.freeze_panes    = "A2"
    step6.auto_filter.ref = f"A1:N{step6.max_row}"

    _build_retail_tab(workbook, step6, headers, processed_rows,
                      total_amount, total_tax, total_total)

    print(f"  [Step6] TOTAL: {time.time()-t0:.2f}s")
    return step6


def _build_retail_tab(workbook, step6, headers, processed_rows,
                      total_amount, total_tax, total_total):
    TAB_NAME = "Retail"

    if TAB_NAME in workbook.sheetnames:
        del workbook[TAB_NAME]

    retail = workbook.create_sheet(TAB_NAME)
    retail.append(headers)

    COL_AMOUNT_OUT = 8
    COL_TAX_OUT    = 9
    COL_TOTAL_OUT  = 10
    COL_NS_CUST_ID_OUT = 14

    for _, full_row, cust_is_yellow in processed_rows:
        retail.append(full_row)
        if cust_is_yellow:
            retail.cell(row=retail.max_row, column=COL_NS_CUST_ID_OUT).fill = DUP_YELLOW

    for row in range(2, retail.max_row + 1):
        for col in [COL_AMOUNT_OUT, COL_TAX_OUT, COL_TOTAL_OUT]:
            cell = retail.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'

    retail.append([])
    retail.append([
        "", "", "", "", "", "", "TOTALS",
        total_amount,
        total_tax,
        total_total,
        "", "", "", ""
    ])

    last_row = retail.max_row
    for col in [COL_AMOUNT_OUT, COL_TAX_OUT, COL_TOTAL_OUT]:
        retail.cell(row=last_row, column=col).number_format = '$#,##0.00'
        retail.cell(row=last_row, column=col).font          = Font(bold=True)
    retail.cell(row=last_row, column=7).font = Font(bold=True)

    for col in range(1, 15):
        retail.column_dimensions[get_column_letter(col)].width = 20

    retail.freeze_panes    = "A2"
    retail.auto_filter.ref = f"A1:N{retail.max_row}"