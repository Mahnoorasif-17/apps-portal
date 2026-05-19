from openpyxl.styles import Alignment, PatternFill

from openpyxl.utils import get_column_letter



import os





ORANGE_FILL = PatternFill(start_color="FFFFD580",

                          end_color="FFFFD580", fill_type="solid")

GRAY_FILL = PatternFill(start_color='DDDDDD',

                        end_color='DDDDDD', fill_type='solid')





class ValidationError(Exception):

    def __init__(self, message, workbook=None):

        super().__init__(message)

        self.workbook = workbook





def generate_new_filename(filepath):
    import os
    import tempfile
    base = os.path.basename(filepath)
    name, ext = os.path.splitext(base)
    # Save to system temp dir — works on both Windows (local) and Linux (Streamlit Cloud)
    return os.path.join(tempfile.gettempdir(), f"{name} - FINAL_PROCESSED{ext}")




def copy_sheet(workbook, source_name, target_name):

    source = workbook[source_name]

    target = workbook.copy_worksheet(source)

    target.title = target_name

    return target





def highlight_row(sheet, row, max_col, fill):

    for col in range(1, max_col + 1):

        sheet.cell(row=row, column=col).fill = fill





def get_footer_row(sheet):
    """Fast: scan from bottom using iter_rows."""
    last_data_row = 1
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        if any(c.value for c in row_cells):
            last_data_row = row_cells[0].row
    return last_data_row




def format_header(sheet, header_row=7):

    regid_col = None

    for col in range(1, sheet.max_column + 1):

        if str(sheet.cell(row=header_row, column=col).value).strip() == "RegID":

            regid_col = col

            break



    if regid_col is None:

        raise ValidationError(f"Expected 'RegID' in row {header_row}")



    sheet.freeze_panes = f"A{header_row + 1}"



    max_col = sheet.max_column

    max_row = sheet.max_row

    for col in range(1, max_col + 1):

        cell = sheet.cell(row=header_row, column=col)

        cell.alignment = Alignment(horizontal='left')

    last_col_letter = get_column_letter(max_col)

    sheet.auto_filter.ref = f"A{header_row}:{last_col_letter}{max_row}"





# ⚡ CHANGED: no longer scans all rows in reverse to find last row

def highlight_rows(sheet, header_row=7):
    """
    Now only colors the header row gray — no longer touches the last row.
    """
    max_col = sheet.max_column
    for col in range(1, max_col + 1):
        sheet.cell(row=header_row, column=col).fill = GRAY_FILL





def highlight_header_row(sheet, header_row=7):

    max_col = sheet.max_column

    for col in range(1, max_col + 1):

        sheet.cell(row=header_row, column=col).fill = GRAY_FILL





def get_column_index_by_header(sheet, header_name, header_row=1):

    for col in range(1, sheet.max_column + 1):

        cell = sheet.cell(row=header_row, column=col)

        if str(cell.value).strip().lower() == header_name.strip().lower():

            return col

    raise ValidationError(

        f"Header '{header_name}' not found in row {header_row}")





def freeze_top_and_filter(sheet):

    sheet.freeze_panes = "A2"

    header_row = 1

    last_col = get_column_letter(sheet.max_column)

    last_row = sheet.max_row

    sheet.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"





def sort_sheet_by_column(sheet, col_index, header_row, last_row):

    data = []

    for row in range(header_row + 1, last_row + 1):

        row_values = [sheet.cell(row=row, column=col).value for col in range(

            1, sheet.max_column + 1)]

        data.append((sheet.cell(row=row, column=col_index).value, row_values))



    data.sort(key=lambda x: (x[0] if x[0] is not None else float('inf')))



    for i, (_, row_values) in enumerate(data, start=header_row + 1):

        for col_idx, value in enumerate(row_values, start=1):

            sheet.cell(row=i, column=col_idx).value = value





# ⚡ CHANGED: collect cols to delete first, then delete — avoids index shifting bugs

def remove_empty_columns(sheet):

    to_delete = []

    for col in range(sheet.max_column, 0, -1):

        col_vals = [sheet.cell(row=row, column=col).value

                    for row in range(1, sheet.max_row + 1)]

        if all(v in (None, "") for v in col_vals):

            to_delete.append(col)

    for col in to_delete:

        sheet.delete_cols(col)





def remove_columns_by_header(sheet, headers):

    header_row = 1

    for header in headers:

        try:

            col = get_column_index_by_header(sheet, header, header_row)

            sheet.delete_cols(col)

        except ValueError:

            continue





# ⚡ CHANGED: collect rows first, delete bottom-up in one pass — avoids repeated shifting

def drop_rows_with_empty_item(sheet):
    import time
    t0 = time.time()
    print(f"    [drop_rows] start, max_row={sheet.max_row}")

    item_col = get_column_index_by_header(sheet, "Item", 1)
    max_col = sheet.max_column

    # Read all rows
    all_rows = []
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col):
        all_rows.append([c.value for c in row_cells])

    header = all_rows[0]
    kept = [header] + [r for r in all_rows[1:] if r[item_col - 1]]

    # Wipe all values in place (no delete_rows, no append)
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col):
        for c in row_cells:
            c.value = None

    # Write kept rows back
    for out_idx, row_values in enumerate(kept, start=1):
        for col_idx, val in enumerate(row_values, start=1):
            sheet.cell(row=out_idx, column=col_idx).value = val

    print(f"    [drop_rows] done in {time.time()-t0:.2f}s, kept={len(kept)}")



def apply_filter_top(sheet):

    last_col = get_column_letter(sheet.max_column)

    sheet.auto_filter.ref = f"A1:{last_col}{sheet.max_row}"





def remove_footer_and_mech_rows(sheet):
    """Fast: find last data row, then wipe everything after it (no delete_rows)."""
    import time
    t0 = time.time()
    print(f"    [remove_footer] start, max_row={sheet.max_row}")

    max_col = sheet.max_column

    # Find last row with any data
    last_data_row = 1
    for row_cells in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=max_col):
        if any(c.value for c in row_cells):
            last_data_row = row_cells[0].row

    print(f"    [remove_footer] last_data_row={last_data_row}")

    # If nothing to remove, done
    if last_data_row >= sheet.max_row:
        print(f"    [remove_footer] nothing to remove, done in {time.time()-t0:.2f}s")
        return

    # Wipe all cells after last_data_row instead of delete_rows
    for row_cells in sheet.iter_rows(min_row=last_data_row + 1, max_row=sheet.max_row, max_col=max_col):
        for c in row_cells:
            c.value = None

    print(f"    [remove_footer] done in {time.time()-t0:.2f}s")




def color_row(sheet, row, fill):

    for col in range(1, sheet.max_column + 1):

        sheet.cell(row=row, column=col).fill = fill





def delete_above_header(sheet):

    for row in sheet.iter_rows(min_row=1, max_row=20):

        if row[0].value == "RegID":

            header_row = row[0].row

            for _ in range(header_row - 1):

                sheet.delete_rows(1)

            return

    raise ValidationError("Header row not found")




from openpyxl.utils import get_column_letter

def autofit_columns(sheet):
    """Lightning fast autofit for yearly files."""
    for col_idx in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        # Only check headers and first 50 rows
        for row_idx in range(1, min(sheet.max_row, 50) + 1):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

_EMPTY_FILL = PatternFill()

def clear_all_highlighting(sheet):
    """Reuses one PatternFill object — fast and correct."""
    for row in sheet.iter_rows():
        for cell in row:
            cell.fill = _EMPTY_FILL

def apply_currency_format(sheet, columns):
    """Applies $#,##0.00 format to specified column indices."""
    for row in range(2, sheet.max_row + 1):
        for col in columns:
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '$#,##0.00'