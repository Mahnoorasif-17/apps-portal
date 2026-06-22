import time
import re
import math
from copy import copy
from collections import defaultdict
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from .utils import *

FILL_PURPLE = PatternFill(start_color="B44CB5", end_color="B44CB5", fill_type="solid")
FILL_BLUE   = PatternFill(start_color="FF0099FF", end_color="FF0099FF", fill_type="solid")
FILL_GREEN  = PatternFill(start_color="FF00CC00", end_color="FF00CC00", fill_type="solid")
FILL_GRAY   = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")


def process_step_5(workbook):
    t0 = time.time()
    step5 = copy_sheet(workbook, "Step 4", "Step 5")
    print(f"  [Step5] copy_sheet: {time.time()-t0:.2f}s (max_row={step5.max_row})")

    item_col     = get_column_index_by_header(step5, "Item", 1)
    customer_col = get_column_index_by_header(step5, "Customer", 1)
    uid_col      = get_column_index_by_header(step5, "UID", 1)
    amount_col   = get_column_index_by_header(step5, "Amount", 1)
    regid_col    = get_column_index_by_header(step5, "RegID", 1)
    max_col      = step5.max_column

    # --- Load entire sheet into memory ONCE ---
    t = time.time()
    rows_data = []
    for row_cells in step5.iter_rows(min_row=1, max_row=step5.max_row, max_col=max_col):
        rows_data.append([c.value for c in row_cells])
    max_row = len(rows_data)
    print(f"  [Step5] loaded {max_row} rows in {time.time()-t:.2f}s")

    # --- Collect UIDs already in service sheets (DHL/USPS/FedEx/UPS) ---
    SERVICE_SHEET_NAMES = ["DHL", "USPS", "FedEx", "UPS"]
    used_uids = set()
    for sname in SERVICE_SHEET_NAMES:
        if sname not in workbook.sheetnames:
            continue
        svc = workbook[sname]
        svc_uid_col = None
        for col in range(1, svc.max_column + 1):
            if str(svc.cell(row=1, column=col).value or "").strip().lower() == "uid":
                svc_uid_col = col
                break
        if svc_uid_col is None:
            continue
        for row_cells in svc.iter_rows(min_row=2, max_row=svc.max_row,
                                        min_col=svc_uid_col, max_col=svc_uid_col):
            v = row_cells[0].value
            if v:
                used_uids.add(v)
    print(f"  [Step5] found {len(used_uids)} UIDs already in service sheets")

    # --- Fix UID number format ---
    t = time.time()
    for r_idx in range(1, max_row):
        v = rows_data[r_idx][uid_col - 1]
        if v:
            step5.cell(row=r_idx + 1, column=uid_col).number_format = '0'
    print(f"  [Step5] UID format: {time.time()-t:.2f}s")

    # --- Coloring classification using in-memory data ---
    MAILBOX_ONLY_KEYWORDS = ["term:", "term "]
    MAILBOX_KEYWORDS      = ["mailbox"]
    RENEW_KEYWORDS        = ["renew"]
    TERM_KEYWORDS         = ["term"]
    SETUP_KEYWORDS        = ["setup fee", "set up fee"]
    INCLUDES_KEYWORDS     = ["includes", "free month"]
    MAILBOX_ROW_KEYWORDS  = ["mailbox", "renew", "term", "setup fee", "set up fee",
                              "includes", "free month", "late fee"]
    MAILBOX_EXCLUSION_KEYWORDS = ["manila", "envelope", "bubble"]

    EXTRA_PURPLE_KEYWORDS = [
        "discount", "coupon", "return", "home depot", "masks black", "error",
        "saran wrap", "lunch", "pay out", "advance", "plumbing",
        "window cleaner", "window washer", "lundh", "pathe", "luis",
        "skyler", "dolly", "chinese"
    ]

    # --- Pre-build: how many rows share each RegID ---
    t = time.time()
    regid_count = defaultdict(int)
    for r_idx in range(1, max_row):
        rid = rows_data[r_idx][regid_col - 1]
        if rid:
            regid_count[rid] += 1
    print(f"  [Step5] regid count map: {time.time()-t:.2f}s")

    t = time.time()
    purple_rows    = set()
    mailbox_regids = set()
    row_color = {}

    for r_idx in range(1, max_row):
        row_data = rows_data[r_idx]
        sheet_row = r_idx + 1

        item_val   = str(row_data[item_col - 1] or "").strip().lower()
        amount_val = row_data[amount_col - 1]
        regid      = row_data[regid_col - 1]

        if "mechanical total" in item_val:
            continue

        if not any(kw in item_val for kw in MAILBOX_ONLY_KEYWORDS):
            is_zero_amount   = (amount_val == 0 or amount_val == 0.0)
            is_petty_cash    = "petty cash" in item_val or "petty cahs" in item_val or "pettycash" in item_val
            is_void          = "void" in item_val
            is_regular_saved = "regular" in item_val and "saved" in item_val
            is_tip           = "tip" in item_val
            is_food_water    = "food" in item_val or "water" in item_val
            is_donation      = "donation" in item_val
            is_petty_pretty  = "petty pretty" in item_val
            is_extra_purple  = any(kw in item_val for kw in EXTRA_PURPLE_KEYWORDS)

            should_be_purple = (
                is_zero_amount or is_petty_cash or is_void or is_regular_saved
                or is_tip or is_food_water or is_donation or is_petty_pretty
                or is_extra_purple
            )

            row_uid = row_data[uid_col - 1]
            if should_be_purple and row_uid and row_uid in used_uids:
                should_be_purple = False

            is_paired_keyword = (
                "coupon" in item_val or
                "discount" in item_val or
                "return" in item_val
            )
            if should_be_purple and is_paired_keyword and regid and regid_count.get(regid, 0) > 1:
                should_be_purple = False

            if should_be_purple:
                purple_rows.add(sheet_row)
                row_color[sheet_row] = FILL_PURPLE

        is_mailbox_item = (
            any(kw in item_val for kw in MAILBOX_KEYWORDS) or
            any(kw in item_val for kw in RENEW_KEYWORDS) or
            any(kw in item_val for kw in SETUP_KEYWORDS) or
            any(kw in item_val for kw in INCLUDES_KEYWORDS)
        )
        if is_mailbox_item and regid:
            mailbox_regids.add(regid)
    print(f"  [Step5] classification pass: {time.time()-t:.2f}s")

    # --- E-Scribers blue ---
    t = time.time()
    for r_idx in range(1, max_row):
        sheet_row = r_idx + 1
        if sheet_row in purple_rows:
            continue
        customer_val = str(rows_data[r_idx][customer_col - 1] or "").strip().lower()
        if "scriber" in customer_val:
            row_color[sheet_row] = FILL_BLUE
    print(f"  [Step5] escriber pass: {time.time()-t:.2f}s")

    # --- regid -> sheet_rows map ---
    t = time.time()
    regid_rows = defaultdict(list)
    for r_idx in range(1, max_row):
        regid = rows_data[r_idx][regid_col - 1]
        if regid in mailbox_regids:
            regid_rows[regid].append(r_idx + 1)
    print(f"  [Step5] regid map: {time.time()-t:.2f}s")

    # --- Collect mailbox rows ---
    t = time.time()
    mailbox_rows = []
    mailbox_row_set = set()

    for regid, sheet_rows in regid_rows.items():
        first_mailbox_idx = None
        for i, sheet_row in enumerate(sheet_rows):
            item_val = str(rows_data[sheet_row - 1][item_col - 1] or "").strip().lower()
            if any(kw in item_val for kw in MAILBOX_ROW_KEYWORDS):
                first_mailbox_idx = i
                break
        if first_mailbox_idx is None:
            continue

        for sheet_row in sheet_rows[first_mailbox_idx:]:
            if sheet_row in mailbox_row_set:
                continue
            item_val = str(rows_data[sheet_row - 1][item_col - 1] or "").strip().lower()
            if any(kw in item_val for kw in MAILBOX_EXCLUSION_KEYWORDS):
                continue
            is_mailbox_row = any(kw in item_val for kw in MAILBOX_ROW_KEYWORDS)
            is_coupon = "coupon" in item_val
            if not is_mailbox_row and not is_coupon:
                continue
            mailbox_rows.append(sheet_row)
            mailbox_row_set.add(sheet_row)
            if not any(kw in item_val for kw in TERM_KEYWORDS):
                if sheet_row not in purple_rows:
                    row_color[sheet_row] = FILL_GREEN
    mailbox_rows.sort()
    print(f"  [Step5] mailbox rows ({len(mailbox_rows)} found): {time.time()-t:.2f}s")

    # --- Apply colors ---
    t = time.time()
    for sheet_row, fill in row_color.items():
        for c in range(1, max_col + 1):
            step5.cell(row=sheet_row, column=c).fill = fill
    print(f"  [Step5] apply colors ({len(row_color)} rows): {time.time()-t:.2f}s")

    # --- Helper column for filtering ---
    t = time.time()
    helper_col = max_col + 1
    helper_col_letter = get_column_letter(helper_col)
    step5.cell(row=1, column=helper_col).value = "_filter"
    for sheet_row in range(2, max_row + 1):
        step5.cell(row=sheet_row, column=helper_col).value = "PURPLE" if sheet_row in purple_rows else "OTHER"
    step5.freeze_panes = "A2"
    step5.auto_filter.ref = f"A1:{get_column_letter(helper_col)}{max_row}"
    step5.column_dimensions[helper_col_letter].hidden = True
    print(f"  [Step5] helper column: {time.time()-t:.2f}s")

    # --- Build downstream sheets ---
    t = time.time()
    _copy_rows_to_tab_fast(rows_data, workbook, "Mailbox", mailbox_rows, max_col, exclude_col=helper_col)
    print(f"  [Step5] Mailbox tab: {time.time()-t:.2f}s")

    t = time.time()
    build_mailbox_working_fast(rows_data, workbook, mailbox_rows, item_col, regid_col,
                                uid_col, customer_col, amount_col,
                                get_column_index_by_header(step5, "Date", 1),
                                get_column_index_by_header(step5, "Time", 1),
                                get_column_index_by_header(step5, "Tender", 1))
    print(f"  [Step5] Mailbox Working: {time.time()-t:.2f}s")

    t = time.time()
    build_void_discount_coupons_fast(rows_data, workbook, purple_rows, item_col, regid_col,
                                       uid_col, customer_col, amount_col,
                                       get_column_index_by_header(step5, "Date", 1),
                                       get_column_index_by_header(step5, "Time", 1),
                                       get_column_index_by_header(step5, "Tender", 1))
    print(f"  [Step5] VDC tab: {time.time()-t:.2f}s")

    t = time.time()
    autofit_columns(step5)
    print(f"  [Step5] autofit: {time.time()-t:.2f}s")

    print(f"  [Step5] TOTAL: {time.time()-t0:.2f}s")


def _copy_rows_to_tab_fast(rows_data, workbook, tab_name, sheet_rows_to_copy, max_col, exclude_col=None):
    if tab_name in workbook.sheetnames:
        del workbook[tab_name]
    ws = workbook.create_sheet(tab_name)

    cols_to_copy = [c for c in range(1, max_col + 1) if c != exclude_col]

    header = rows_data[0]
    for out_col, src_col in enumerate(cols_to_copy, start=1):
        ws.cell(row=1, column=out_col).value = header[src_col - 1]

    write_row = 2
    for sheet_row in sheet_rows_to_copy:
        src = rows_data[sheet_row - 1]
        for out_col, src_col in enumerate(cols_to_copy, start=1):
            cell = ws.cell(row=write_row, column=out_col)
            cell.value = src[src_col - 1]
            cell.fill = FILL_GREEN
        write_row += 1

    freeze_top_and_filter(ws)
    highlight_header_row(ws, header_row=1)
    autofit_columns(ws)


def build_void_discount_coupons_fast(rows_data, workbook, purple_rows,
                                       item_col, regid_col, uid_col, customer_col,
                                       amount_col, date_col, time_col, tender_col):
    TAB_NAME = "Void-Discount-Coupons"
    if TAB_NAME in workbook.sheetnames:
        del workbook[TAB_NAME]
    ws = workbook.create_sheet(TAB_NAME)

    COL_UID, COL_REGID, COL_DATE, COL_TIME = 1, 2, 3, 4
    COL_ITEM, COL_TENDER, COL_CUSTOMER, COL_AMOUNT = 5, 6, 7, 8

    headers = ["UID", "RegID", "Date", "Time", "Item", "Tender", "Customer", "Amount"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = FILL_PURPLE
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(8)}1"

    col_map = {COL_UID: uid_col, COL_REGID: regid_col, COL_DATE: date_col,
               COL_TIME: time_col, COL_ITEM: item_col, COL_TENDER: tender_col,
               COL_CUSTOMER: customer_col, COL_AMOUNT: amount_col}

    write_row = 2
    for sheet_row in sorted(purple_rows):
        src = rows_data[sheet_row - 1]

        item_check = str(src[item_col - 1] or "").lower()
        if "mechanical total" in item_check:
            continue

        for out_col, src_col in col_map.items():
            cell = ws.cell(row=write_row, column=out_col)
            cell.value = src[src_col - 1]
            cell.fill = FILL_PURPLE
            if out_col == COL_UID:
                cell.number_format = '0'
            elif out_col == COL_AMOUNT:
                cell.number_format = '$#,##0.00'
                v = src[src_col - 1]
                if isinstance(v, (int, float)) and v < 0:
                    cell.font = Font(color="FF0000")
            elif out_col == COL_DATE:
                cell.number_format = 'mm/dd/yyyy'
        write_row += 1

    autofit_columns(ws)


def build_mailbox_working_fast(rows_data, workbook, mailbox_rows,
                                 item_col, regid_col, uid_col, customer_col,
                                 amount_col, date_col, time_col, tender_col):
    TAB_NAME = "Mailbox Working"
    if TAB_NAME in workbook.sheetnames:
        del workbook[TAB_NAME]
    ws = workbook.create_sheet(TAB_NAME)

    COL_UID, COL_REGID, COL_DATE, COL_TIME = 1, 2, 3, 4
    COL_ITEM, COL_MBOX_NUM, COL_MBOX_TYP = 5, 6, 7
    COL_TENDER, COL_CUSTOMER, COL_AMOUNT, COL_TAX, COL_TOTAL = 9, 10, 11, 12, 13

    headers = {COL_UID: "UID", COL_REGID: "RegID", COL_DATE: "Date", COL_TIME: "Time",
               COL_ITEM: "Item", COL_MBOX_NUM: "Mailbox #", COL_MBOX_TYP: "Mailbox Type",
               COL_TENDER: "Tender", COL_CUSTOMER: "Customer",
               COL_AMOUNT: "Amount", COL_TAX: "Tax", COL_TOTAL: "Total Amount"}
    for col, h in headers.items():
        ws.cell(row=1, column=col).value = h

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(13)}1"
    highlight_header_row(ws, header_row=1)

    def extract_mailbox_number(item_text):
        if not item_text:
            return None
        m = re.search(r'mailbox\s*#(\d+)', item_text, re.IGNORECASE)
        if m:
            return int(m.group(1))
        m = re.search(r'mailbox\s+(\d+)', item_text, re.IGNORECASE)
        if m:
            return int(m.group(1))
        return None

    def extract_mailbox_type(item_text):
        if item_text and "business" in item_text.lower():
            return " BUSINESS"
        return None

    def is_term_or_child_row(item_text):
        il = (item_text or "").strip().lower()
        return il.startswith("term") or "  term" in il or il.startswith("term:")

    # ---- Block detection: a "block" = parent Renew/Setup/Mailbox + its children ----
    PARENT_KEYWORDS = ["renew", "mailbox", "setup fee", "set up fee"]

    def is_parent_row(item_text):
        il = (item_text or "").strip().lower()
        if il.startswith("term") or "  term" in il:
            return False
        return any(kw in il for kw in PARENT_KEYWORDS)

    blocks_by_parent = []
    current_parent = None
    current_children = []

    for sheet_row in mailbox_rows:
        src = rows_data[sheet_row - 1]
        item_val_scan = str(src[item_col - 1] or "")
        if is_parent_row(item_val_scan):
            if current_parent is not None:
                blocks_by_parent.append((current_parent, current_children))
            current_parent = sheet_row
            current_children = []
        else:
            if current_parent is not None:
                current_children.append(sheet_row)
            else:
                blocks_by_parent.append((sheet_row, []))

    if current_parent is not None:
        blocks_by_parent.append((current_parent, current_children))

    # Mark zero-tax rows: any block that contains a coupon
    zero_tax_rows = set()
    for parent_row, child_rows in blocks_by_parent:
        block_all_rows = [parent_row] + child_rows
        has_coupon = False
        for r in block_all_rows:
            src = rows_data[r - 1]
            item_val_scan = str(src[item_col - 1] or "").lower()
            if "coupon" in item_val_scan:
                has_coupon = True
                break
        if has_coupon:
            for r in block_all_rows:
                zero_tax_rows.add(r)

    # Debug print so you can verify in terminal
    print(f"  [MailboxWorking] {len(blocks_by_parent)} blocks, {len(zero_tax_rows)} zero-tax rows")

    write_row = 2
    last_mbox_row_for_regid = {}

    for sheet_row in mailbox_rows:
        src = rows_data[sheet_row - 1]
        item_val = str(src[item_col - 1] or "")
        amount_val = src[amount_col - 1] or 0
        regid_val = src[regid_col - 1]

        # Mailbox #
        mbox_num = extract_mailbox_number(item_val)
        if mbox_num is not None:
            ws.cell(row=write_row, column=COL_MBOX_NUM).value = mbox_num
            last_mbox_row_for_regid[regid_val] = write_row
        else:
            parent_row = last_mbox_row_for_regid.get(regid_val)
            if parent_row:
                ws.cell(row=write_row, column=COL_MBOX_NUM).value = f"=F{parent_row}"

        # Mailbox Type
        mbox_type = extract_mailbox_type(item_val)
        if mbox_type:
            ws.cell(row=write_row, column=COL_MBOX_TYP).value = mbox_type
            last_mbox_row_for_regid[str(regid_val) + "_type"] = mbox_type
        elif is_term_or_child_row(item_val):
            inherited = last_mbox_row_for_regid.get(str(regid_val) + "_type")
            ws.cell(row=write_row, column=COL_MBOX_TYP).value = inherited

        # Core fields
        ws.cell(row=write_row, column=COL_UID).value      = src[uid_col - 1]
        ws.cell(row=write_row, column=COL_REGID).value    = src[regid_col - 1]
        ws.cell(row=write_row, column=COL_DATE).value     = src[date_col - 1]
        ws.cell(row=write_row, column=COL_TIME).value     = src[time_col - 1]
        ws.cell(row=write_row, column=COL_ITEM).value     = item_val
        ws.cell(row=write_row, column=COL_TENDER).value   = src[tender_col - 1]
        ws.cell(row=write_row, column=COL_CUSTOMER).value = src[customer_col - 1]
        ws.cell(row=write_row, column=COL_AMOUNT).value   = amount_val

        # --- Tax: COMPUTED in Python (not formula) for reliability ---
       # --- Tax: COMPUTED in Python with round-half-up (so 37.275 -> 37.28) ---
        item_val_lower = item_val.lower()
        is_late_fee = "late fee" in item_val_lower

        if sheet_row in zero_tax_rows or is_late_fee:
            tax_value = 0.0
        else:
            if isinstance(amount_val, (int, float)):
                # Round HALF UP (not banker's): 37.275 -> 37.28
                tax_raw = amount_val * 0.08875
                tax_value = math.floor(tax_raw * 100 + 0.5) / 100
            else:
                tax_value = 0.0
        ws.cell(row=write_row, column=COL_TAX).value = tax_value

        # --- Total: COMPUTED in Python (same rounding) ---
        if isinstance(amount_val, (int, float)):
            total_raw = amount_val + tax_value
            total_value = math.floor(total_raw * 100 + 0.5) / 100 if total_raw >= 0 else -math.floor(-total_raw * 100 + 0.5) / 100
        else:
            total_value = tax_value
        ws.cell(row=write_row, column=COL_TOTAL).value = total_value

        # Formats
        ws.cell(row=write_row, column=COL_UID).number_format    = '0'
        ws.cell(row=write_row, column=COL_AMOUNT).number_format = '$#,##0.00'
        ws.cell(row=write_row, column=COL_TAX).number_format    = '$#,##0.00'
        ws.cell(row=write_row, column=COL_TOTAL).number_format  = '$#,##0.00'
        ws.cell(row=write_row, column=COL_DATE).number_format   = 'mm/dd/yyyy'

        # Color green
        for col in range(1, 14):
            ws.cell(row=write_row, column=col).fill = FILL_GREEN

        write_row += 1

    # --- Totals row: COMPUTED from written values ---
    last_data_row = write_row - 1
    totals_row = write_row + 1

    sum_amount = 0.0
    sum_tax    = 0.0
    sum_total  = 0.0
    for r in range(2, write_row):
        a = ws.cell(row=r, column=COL_AMOUNT).value
        t = ws.cell(row=r, column=COL_TAX).value
        tot = ws.cell(row=r, column=COL_TOTAL).value
        if isinstance(a, (int, float)):
            sum_amount += a
        if isinstance(t, (int, float)):
            sum_tax += t
        if isinstance(tot, (int, float)):
            sum_total += tot

    def half_up(val):
        if val >= 0:
            return math.floor(val * 100 + 0.5) / 100
        else:
            return -math.floor(-val * 100 + 0.5) / 100

    ws.cell(row=totals_row, column=COL_AMOUNT).value = half_up(sum_amount)
    ws.cell(row=totals_row, column=COL_TAX).value    = half_up(sum_tax)
    ws.cell(row=totals_row, column=COL_TOTAL).value  = half_up(sum_total)

    for c in (COL_AMOUNT, COL_TAX, COL_TOTAL):
        ws.cell(row=totals_row, column=c).number_format = '$#,##0.00'
        ws.cell(row=totals_row, column=c).font = Font(bold=True)
    ws.cell(row=totals_row, column=COL_CUSTOMER).value = "Total:"
    ws.cell(row=totals_row, column=COL_CUSTOMER).font = Font(bold=True)

    autofit_columns(ws)