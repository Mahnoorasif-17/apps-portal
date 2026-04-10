from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from .utils import *

ORANGE_FILL = PatternFill(start_color="FFFFD580",
                          end_color="FFFFD580", fill_type="solid")


def process_step_2(workbook):
    step2_sheet = copy_sheet(workbook, "Step 1", "Step 2")

    delete_above_header(step2_sheet)
    fix_column_headers_for_sheet_2(step2_sheet)
    format_header(step2_sheet, header_row=1)
    footer_row = get_footer_row(step2_sheet)

    col_totals = insert_mechanical_totals(step2_sheet, footer_row)
    validate_totals(step2_sheet, footer_row, col_totals)
    sum_formula_row = adjust_amount_total_with_deductions(
        step2_sheet, footer_row)
    finalize_adjusted_total_validation(step2_sheet, sum_formula_row)


def insert_mechanical_totals(sheet, footer_row):
    sheet.insert_rows(footer_row, 2)
    mech_row = footer_row
    sheet[f"C{mech_row}"] = "Mechanical Totals"

    headers = ["Amount", "SubTotal", "Tax", "Total"]
    col_totals = {}

    for header in headers:
        col_index = get_column_index_by_header(sheet, header, header_row=1)
        col_letter = get_column_letter(col_index)

        start, end = 2, mech_row - 1
        formula = f"=SUM({col_letter}{start}:{col_letter}{end})"
        cell = sheet.cell(row=mech_row, column=col_index)
        cell.value = formula
        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        # Calculate total using values directly
        total = 0
        for row in sheet.iter_rows(min_row=start, max_row=end, min_col=col_index, max_col=col_index):
            val = row[0].value
            if isinstance(val, (int, float)):
                total += val

        col_totals[col_letter] = total

    return col_totals


def validate_totals(sheet, footer_row, col_totals):
    headers = ["SubTotal", "Tax", "Total"]
    for header in headers:
        col_index = get_column_index_by_header(sheet, header, header_row=1)
        col_letter = get_column_letter(col_index)
        footer_val = sheet.cell(row=footer_row + 2, column=col_index).value
        calc_val = col_totals[col_letter]

        if footer_val is not None and not str(footer_val).startswith("="):
            try:
                clean = str(footer_val).replace(
                    "(", "-").replace(")", "").replace("$", "").replace(",", "")
                footer_float = float(clean)
                if round(calc_val, 2) != round(footer_float, 2):
                    raise ValidationError(
                        f"Transaction Count's SubTotal does not match Mechanical Total's SubTotal.\n"
                        f"Mismatch in column '{header}': {calc_val:.2f} ≠ {footer_float:.2f}"
                    )
            except Exception:
                raise ValidationError(
                    f"Transaction Count's SubTotal does not match Mechanical Total's SubTotal. (row {footer_row + 2})"
                )


def adjust_amount_total_with_deductions(sheet, mech_row):
    col_item = get_column_index_by_header(sheet, "Item")
    col_amount = get_column_index_by_header(sheet, "Amount")
    col_subtotal = get_column_index_by_header(sheet, "SubTotal")

    # Step 1: Adjust Amount based on negative SubTotal
    for row in range(2, mech_row):
        sub_val = sheet.cell(row=row, column=col_subtotal).value
        if isinstance(sub_val, (int, float)) and sub_val < 0:
            next_row = row + 1
            amt_cell = sheet.cell(row=next_row, column=col_amount)
            amt_val = amt_cell.value
            if isinstance(amt_val, (int, float)):
                if amt_val > 0:
                    amt_cell.value = -amt_val
                highlight_row(sheet, row, sheet.max_column, ORANGE_FILL)

    # Step 2: Adjust Amount if Item contains keyword (unless it says VOID)
    keywords = ["discount", "coupon", "petty"]
    for row in range(2, mech_row):
        item = sheet.cell(row=row, column=col_item).value
        if isinstance(item, str):
            lower = item.lower()
            if "void" in lower:
                continue
            if any(kw in lower for kw in keywords):
                amt_cell = sheet.cell(row=row, column=col_amount)
                amt_val = amt_cell.value
                if isinstance(amt_val, (int, float)) and amt_val > 0:
                    amt_cell.value = -amt_val
                if isinstance(amt_val, (int, float)):
                    highlight_row(sheet, row, sheet.max_column, ORANGE_FILL)

    # Step 3: Insert Difference Formula Row (in Excel)
    difference_row = mech_row + 1
    sheet[f"B{difference_row}"] = "Difference"

    footer_row = get_footer_row(sheet)

    for header in ["Amount", "SubTotal", "Tax", "Total"]:
        col_index = get_column_index_by_header(sheet, header)
        col_letter = get_column_letter(col_index)

        diff_formula = f"={col_letter}{mech_row}-{col_letter}{footer_row}"
        cell = sheet.cell(row=difference_row, column=col_index)
        cell.value = diff_formula
        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    return difference_row


def finalize_adjusted_total_validation(sheet, mech_row):
    data_start = 2
    data_end = mech_row - 1

    amount_col = get_column_index_by_header(sheet, "Amount")
    subtotal_col = get_column_index_by_header(sheet, "SubTotal")

    amount_sum = compute_column_sum_by_index(
        sheet, amount_col, data_start, data_end)
    subtotal_sum = compute_column_sum_by_index(
        sheet, subtotal_col, data_start, data_end)

    sheet.cell(row=mech_row + 1, column=amount_col).value = amount_sum
    sheet.cell(row=mech_row + 1,
               column=amount_col).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    if round(amount_sum, 2) != round(subtotal_sum, 2):
        raise ValidationError(
            f"Transaction Count's SubTotal does not match Mechanical Total's SubTotal ({amount_sum:.2f} != {subtotal_sum:.2f})."
        )


def compute_column_sum_by_index(sheet, col_index, start_row, end_row):
    total = 0
    for row in range(start_row, end_row + 1):
        val = sheet.cell(row=row, column=col_index).value
        if isinstance(val, (int, float)):
            total += val
    return total


def fix_column_headers_for_sheet_2(sheet):
    if not sheet["C1"].value:
        sheet["C1"] = "Item"
    if not sheet["H1"].value:
        sheet["H1"] = "Amount"
