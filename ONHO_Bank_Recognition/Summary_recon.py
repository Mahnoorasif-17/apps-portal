import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import numbers as numfmt
from openpyxl.utils import get_column_letter

# =========================
# Helpers
# =========================
def clean_headers(cols):
    # strip and remove non-breaking spaces
    return cols.str.replace("\u00a0", "", regex=False).str.strip()


def clean_object_columns(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = (
                df[c]
                .astype(str)
                .str.replace("\u00a0", "", regex=False)
                .str.strip()
            )
    return df

def write_currency(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    cell.value = float(value) if value is not None else 0.0
    cell.number_format = numfmt.FORMAT_CURRENCY_USD_SIMPLE
    return cell

def autofit_columns(ws):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for c in col_cells:
            v = c.value
            if v is None:
                continue
            if isinstance(v, (int, float)) and ("$" in (c.number_format or "") or "#,##0.00" in (c.number_format or "")):
                disp = f"${abs(v):,.2f}" if v < 0 else f"${v:,.2f}"
            else:
                disp = str(v)
            if len(disp) > max_len:
                max_len = len(disp)
        ws.column_dimensions[col_letter].width = max_len + 2

def pick_col(df, candidates):
    """Return the first candidate that exists in df.columns (exact match), else None."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


# =========================
# Main function
# =========================
def run_reconciliation(batch_file, rta_file, output_file=None):
    # =========================
    # Load files
    # =========================
    batch_df = pd.read_excel(batch_file)
    rta_df   = pd.read_excel(rta_file)

    # =========================
    # Clean headers & values
    # =========================
    batch_df.columns = clean_headers(batch_df.columns)
    rta_df.columns   = clean_headers(rta_df.columns)

    batch_df = clean_object_columns(batch_df)
    rta_df   = clean_object_columns(rta_df)

    # =========================
    # Parse numeric amounts (BATCH made robust like RTA)
    # =========================
    batch_amount_col = pick_col(batch_df, ["Amount", "Total", "Total Amount", "Grand Total", "Amt"])
    if batch_amount_col is None:
        raise ValueError(f"Could not find Amount column in Batch. Columns: {list(batch_df.columns)}")

    batch_df["Amount"] = pd.to_numeric(
        batch_df[batch_amount_col].replace(r"[\$,]", "", regex=True), errors="coerce"
    ).round(2)

    # --- RTA total (already robust) ---
    total_col = pick_col(rta_df, ["Total", "Amount", "Total Amount", "Grand Total", "Amt"])
    if total_col is None:
        raise ValueError(f"Could not find Total/Amount column in RTA. Columns: {list(rta_df.columns)}")

    rta_df["Total"] = pd.to_numeric(
        rta_df[total_col].replace(r"[\$,]", "", regex=True), errors="coerce"
    ).round(2)
    

    # =========================
    # Dates (BATCH made robust like RTA)
    # =========================
    batch_date_col = pick_col(batch_df, ["Batch Date", "Date", "Transaction Date", "Batch_Date", "BatchDate"])
    if batch_date_col is None:
        raise ValueError(f"Could not find a Date column in Batch file. Columns: {list(batch_df.columns)}")

    batch_df["Batch Date"] = pd.to_datetime(batch_df[batch_date_col], errors="coerce")
    batch_df = batch_df.dropna(subset=["Batch Date"]).copy()
    if batch_df.empty:
        raise ValueError(
            "No parsable dates in Batch after parsing possible date columns "
            f"(tried: 'Batch Date', 'Date', 'Transaction Date', 'Batch_Date', 'BatchDate')."
        )

    # We will use the range of batch dates for filtering RTA
    batch_min = batch_df["Batch Date"].dt.date.min()
    batch_max = batch_df["Batch Date"].dt.date.max()
    # Use the first date only to name the output file (keeps your old convention)
    target_date = batch_df["Batch Date"].dt.date.sort_values().iloc[0]

    # Pretty range for headers
    if batch_min == batch_max:
        try:
            summary_date_range_str = batch_min.strftime("%#m/%#d/%Y")
        except Exception:
            summary_date_range_str = batch_min.strftime("%m/%d/%Y")
    else:
        try:
            summary_date_range_str = f"{batch_min.strftime('%#m/%#d/%Y')} - {batch_max.strftime('%#m/%#d/%Y')}"
        except Exception:
            summary_date_range_str = f"{batch_min.strftime('%m/%d/%Y')} - {batch_max.strftime('%m/%d/%Y')}"

    # Robust RTA datetime column
    dt_col = pick_col(rta_df, ["Date/Time", "Date Time", "Datetime", "Date", "Transaction Date"])
    if dt_col is None:
        raise ValueError(f"Could not find a Date/Time column in RTA. Columns: {list(rta_df.columns)}")

    rta_df["__DateTime"] = pd.to_datetime(rta_df[dt_col], errors="coerce")
    rta_df["Date"] = rta_df["__DateTime"].dt.date

    # Filter RTA to batch range; if empty, fall back to all rows
    rta_before = len(rta_df)
    rta_df_filt = rta_df[rta_df["Date"].between(batch_min, batch_max)]
    if rta_df_filt.empty:
        print(f"⚠️ RTA filter by {batch_min}..{batch_max} returned 0 rows; using ALL RTA rows.")
        rta_df_filt = rta_df.copy()
    rta_df = rta_df_filt.reset_index(drop=True)

    # =========================
    # Normalize Tender / Brand (BATCH made robust like RTA)
    # =========================
    # RTA Tender
    tender_src = "Tender" if "Tender" in rta_df.columns else pick_col(
        rta_df, ["Card / Tender", "Card/Tender", "Card Tender", "Card brand", "Card"]
    )
    if tender_src is None:
        raise ValueError(f"Could not find Tender column in RTA. Columns: {list(rta_df.columns)}")

    rta_df["Tender"] = (
        rta_df[tender_src]
        .astype(str).str.replace("\u00a0", "", regex=False).str.lower().str.strip()
        .replace({
            "amex": "american express", "americanexpress": "american express", "american express": "american express",
            "mstrcard": "mastercard", "mc": "mastercard", "master card": "mastercard", "mastercard": "mastercard",
            "visa": "visa", "cash": "cash", "other": "other", "discover": "discover",
            "split": "split", "check": "check", "cheque": "check", "account": "account",
        })
    )

    # BATCH Card brand (robust source)
    brand_src = "Card brand" if "Card brand" in batch_df.columns else pick_col(
        batch_df, ["Card / Tender", "Card/Tender", "Card Tender", "Card", "Tender", "Card Type", "Brand"]
    )
    if brand_src is None:
        raise ValueError(f"Could not find card brand/tender column in Batch. Columns: {list(batch_df.columns)}")

    batch_df["Card brand"] = (
        batch_df[brand_src]
        .astype(str).str.replace("\u00a0", "", regex=False).str.lower().str.strip()
        .replace({
            "amex": "american express", "americanexpress": "american express", "american express": "american express",
            "mstrcard": "mastercard", "master card": "mastercard", "mastercard": "mastercard",
            "visa": "visa", "cash": "cash", "other": "other", "discover": "discover",
            "split": "split", "check": "check", "cheque": "check", "account": "account",
        })
    )

    # =========================
    # Summaries (added split/check/account into modes)
    # =========================
    batch_modes = ["american express", "mastercard", "visa", "discover", "other", "split", "check", "account"]
    all_modes   = batch_modes + ["cash"]

    batch_summary = (
        batch_df.groupby("Card brand", dropna=False)["Amount"].sum()
        .reindex(batch_modes, fill_value=0)
    )

    rta_summary = (
        rta_df.groupby("Tender", dropna=False)["Total"].sum()
        .reindex(all_modes, fill_value=0)
    )

    merged = pd.DataFrame(index=all_modes)
    merged["Amount_Bank"] = batch_summary.reindex(all_modes, fill_value=0)
    merged["Amount_RTA"]  = rta_summary
    merged["Diff"]        = merged["Amount_Bank"] - merged["Amount_RTA"]

    diff_exc_cash = merged.drop("cash").Diff.sum()
    cash_diff     = -merged.loc["cash", "Amount_RTA"] if "cash" in merged.index else 0.0
    total_diff    = diff_exc_cash + cash_diff
    batch_total   = merged.loc[batch_modes, "Amount_Bank"].sum()
    rta_total     = merged.loc[all_modes, "Amount_RTA"].sum()

    # --- Diagnostics ---
    print(f"Batch rows: {len(batch_df)}, date range: {batch_min}..{batch_max}")
    print(f"RTA rows (kept): {len(rta_df)} (from {rta_before})")
    print("RTA per-tender totals used:")
    print(rta_summary.to_string())
    print()

    # =========================
    # Matching for unmatched tables
    # =========================
    # --- Initial Matching (to determine 'Matched' flag based on Rule 1: Date, Tender, Amount) ---
    
    # BATCH to RTA matching
    rta_copy = rta_df.copy()
    matched_flags_batch = []
    
    # NEW logic: List to capture exact matches for the 2nd tab
    exact_matches_for_tab2 = []
    rta_regid_col_orig = pick_col(rta_df, ["RegID", "Reg ID", "ID"])
    batch_cardnum_col_orig = pick_col(batch_df, ["Card number", "Customer", "Customer Name", "Name", "Account", "Masked Card", "Last 4", "Last4"])

    for index, batch_row in batch_df.iterrows():
        b_amount = batch_row["Amount"]
        b_brand  = batch_row["Card brand"]
        b_date   = batch_row["Batch Date"].date() # Get date object for batch
        
        # Rule 1 Implementation: Match Date, Amount and Tender/Card brand
        match = rta_copy[(rta_copy["Total"] == b_amount) & 
                         (rta_copy["Tender"] == b_brand) &
                         (rta_copy["Date"] == b_date)] 

        if not match.empty:
            match_row = match.iloc[0]
            # Save for Tab 2
            exact_matches_for_tab2.append({
                "Batch Date": batch_row["Batch Date"],
                "Card brand": b_brand.capitalize(),
                "Card number": batch_row[batch_cardnum_col_orig] if batch_cardnum_col_orig else "",
                "Amount": b_amount,
                "RTA RegID": match_row[rta_regid_col_orig] if rta_regid_col_orig else "",
                "RTA Date/Time": match_row["__DateTime"]
            })
            
            rta_copy = rta_copy.drop(index=match.index[0])
            matched_flags_batch.append(True)
        else:
            matched_flags_batch.append(False)
    batch_df["Matched"] = matched_flags_batch
    unmatched_batch = batch_df[~batch_df["Matched"]].copy()

    # Robust "Card number" source for Batch (so table1 never KeyErrors)
    batch_cardnum_col = pick_col(unmatched_batch, ["Card number", "Customer", "Customer Name", "Name", "Account", "Masked Card", "Last 4", "Last4"])
    table1 = pd.DataFrame({
        "Batch Date": unmatched_batch["Batch Date"],
        "Card brand": unmatched_batch["Card brand"],
        "Card number": unmatched_batch[batch_cardnum_col] if batch_cardnum_col else "",
        "Amount": unmatched_batch["Amount"],
    })

    # RTA to BATCH matching
    batch_copy = batch_df.copy()
    matched_flags_rta = []
    for index, rta_row in rta_df.iterrows():
        r_amount = rta_row["Total"]
        r_tender = rta_row["Tender"]
        r_date   = rta_row["Date"] # RTA 'Date' is already date object

        # Rule 1 Implementation: Match Date, Amount and Tender/Card brand
        match = batch_copy[(batch_copy["Amount"] == r_amount) & 
                           (batch_copy["Card brand"] == r_tender) &
                           (batch_copy["Batch Date"].dt.date == r_date)] 

        if not match.empty:
            batch_copy = batch_copy.drop(index=match.index[0])
            matched_flags_rta.append(True)
        else:
            matched_flags_rta.append(False)
    rta_df["Matched"] = matched_flags_rta
    unmatched_rta = rta_df[~rta_df["Matched"]].copy()

    # ---------- Robust TABLE2 build (for RTA unmatched - now includes RegID) ----------
    cand_datetime = ["Date/Time", "Date Time", "Datetime", "Date", "__DateTime"]
    cand_tender   = ["Tender", "Card / Tender", "Card/Tender", "Card Tender", "Card brand", "Card"]
    cand_customer = ["Customer", "Customer Name", "Name", "Card number", "Account"]
    cand_total    = ["Total", "Amount", "Total Amount", "Grand Total", "Amt"]
    cand_regid    = ["RegID", "Reg ID", "ID"] 

    c_dt = pick_col(unmatched_rta, cand_datetime)
    c_td = pick_col(unmatched_rta, cand_tender)
    c_cu = pick_col(unmatched_rta, cand_customer)
    c_to = pick_col(unmatched_rta, cand_total)
    c_ri = pick_col(unmatched_rta, cand_regid) 

    n = len(unmatched_rta)
    table2_cols = {}

    # Date/Time
    if c_dt:
        table2_cols["Date/Time"] = unmatched_rta[c_dt]
    else:
        table2_cols["Date/Time"] = pd.Series([pd.NaT] * n)

    # RegID
    if c_ri:
        table2_cols["RegID"] = unmatched_rta[c_ri]
    else:
        table2_cols["RegID"] = pd.Series([""] * n)

    # Card brand
    if c_td:
        table2_cols["Card brand"] = unmatched_rta[c_td].astype(str).str.lower().str.strip()
    else:
        table2_cols["Card brand"] = pd.Series([""] * n)

    # Card number
    if c_cu:
        table2_cols["Card number"] = unmatched_rta[c_cu]
    else:
        table2_cols["Card number"] = pd.Series([""] * n)

    # Amount
    if c_to:
        amt = unmatched_rta[c_to]
        if amt.dtype == object:
            amt = amt.replace(r"[\$,]", "", regex=True)
        table2_cols["Amount"] = pd.to_numeric(amt, errors="coerce").round(2)
    else:
        table2_cols["Amount"] = pd.Series([0.0] * n)
    
    # Explicitly set the column order
    table2_columns_order = ["Date/Time", "RegID", "Card brand", "Card number", "Amount"]
    table2 = pd.DataFrame(table2_cols)[table2_columns_order]
    # -----------------------------------------------------

    # --- Secondary Matching (for unmatched transactions: Rule 1 & Rule 2) ---
    batch_unmatched = table1.copy().reset_index(drop=True)
    rta_unmatched   = table2.copy().reset_index(drop=True)
    
    # Add Date columns for easy comparison in the loop
    batch_unmatched["Date"] = batch_unmatched["Batch Date"].dt.date
    rta_unmatched["Date"] = pd.to_datetime(rta_unmatched["Date/Time"], errors='coerce').dt.date 

    batch_unmatched["Matching"] = ""
    batch_unmatched["Comments"] = ""
    rta_unmatched["Matching"]   = ""
    rta_unmatched["Comments"]   = ""

    match_id = 1
    used_rta = set()
    priority = ["american express", "visa", "mastercard", "discover", "split", "check", "account", "other", "cash"]

    for i, b_row in batch_unmatched.iterrows():
        b_amt = b_row["Amount"]
        b_brand = b_row["Card brand"]
        b_date = b_row["Date"] 
        
        # Rule 2: Find candidates matching on Amount and Date (Brand can differ)
        candidates = rta_unmatched[
            (~rta_unmatched.index.isin(used_rta)) & 
            (rta_unmatched["Amount"] == b_amt) &
            (rta_unmatched["Date"] == b_date)
        ]
        
        # --- Rule 1: Same Date, Tender (Card Brand), AND Amount (Priority Match on unmatched pool) ---
        exact_match = candidates[candidates["Card brand"] == b_brand]
        
        if not exact_match.empty:
            r_idx = exact_match.index[0]
            batch_unmatched.at[i, "Matching"] = match_id
            rta_unmatched.at[r_idx, "Matching"] = match_id
            match_id += 1
            used_rta.add(r_idx)
            continue
            
        # --- Continue with existing categorization logic on candidates (Rule 2 covers Date+Amount match) ---

        # 2. Categorization difference match (Batch brand is 'other', RTA brand is card)
        card_match = candidates[candidates["Card brand"].isin(["visa", "mastercard", "american express", "discover", "split", "check", "account"])]
        if not card_match.empty and b_brand == "other":
            r_idx = card_match.index[0]
            r_brand = rta_unmatched.at[r_idx, "Card brand"]
            batch_unmatched.at[i, "Matching"] = match_id
            rta_unmatched.at[r_idx, "Matching"] = match_id
            batch_unmatched.at[i, "Comments"] = f"Categorized as {r_brand.capitalize()} in RTA"
            rta_unmatched.at[r_idx, "Comments"] = f"Categorize as Other in 1881"
            match_id += 1
            used_rta.add(r_idx)
            continue
            
        # 3. Fallback match (Date and Amount match, but brands are different/weird)
        fallback = candidates.sort_values(by="Card brand", key=lambda x: x.map(lambda val: priority.index(val) if val in priority else 999))
        if not fallback.empty:
            r_idx = fallback.index[0]
            r_brand = rta_unmatched.at[r_idx, "Card brand"]
            batch_unmatched.at[i, "Matching"] = match_id
            rta_unmatched.at[r_idx, "Matching"] = match_id
            rta_unmatched.at[r_idx, "Comments"] = f"Categorize as {b_brand.capitalize()} in 1881"
            batch_unmatched.at[i, "Comments"] = f"Matched with {r_brand.capitalize()} in RTA"
            match_id += 1
            used_rta.add(r_idx)

    # Remove the temporary 'Date' column before writing to Excel
    batch_unmatched = batch_unmatched.drop(columns=["Date"])
    rta_unmatched = rta_unmatched.drop(columns=["Date"])


    # =========================
    # Excel output
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Recon Summary"

    bold = Font(bold=True)
    red = Font(color="FF0000", bold=True)
    blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Title
    ws.append(["Summary Recon"])
    ws["A1"].font = bold
    ws.append([])

    # 1881 summary (DATE RANGE just for display)
    ws.append(["1881", summary_date_range_str])
    for col in ["A", "B"]:
        ws[f"{col}{ws.max_row}"].font = bold
        ws[f"{col}{ws.max_row}"].fill = blue_fill
    for mode in batch_modes:
        ws.append([mode.capitalize(), None])
        write_currency(ws, ws.max_row, 2, merged.loc[mode, 'Amount_Bank'])
    ws.append([])

    # RTA summary (DATE RANGE just for display)
    ws.append(["RTA", summary_date_range_str])
    for col in ["A", "B"]:
        ws[f"{col}{ws.max_row}"].font = bold
        ws[f"{col}{ws.max_row}"].fill = blue_fill
    for mode in all_modes:
        ws.append([mode.capitalize(), None])
        write_currency(ws, ws.max_row, 2, merged.loc[mode, 'Amount_RTA'])
    ws.append([])

    # Diff summary
    ws.append(["Diff", summary_date_range_str])
    for col in ["A", "B"]:
        ws[f"{col}{ws.max_row}"].font = bold
        ws[f"{col}{ws.max_row}"].fill = blue_fill
    for mode in batch_modes:
        ws.append([mode.capitalize(), None])
        write_currency(ws, ws.max_row, 2, merged.loc[mode, 'Diff'])
    ws.append(["Diff exc. Cash", None])
    write_currency(ws, ws.max_row, 2, diff_exc_cash)

    ws.append(["Cash", None])
    cash_row = ws.max_row
    write_currency(ws, cash_row, 2, merged.loc["cash", "Amount_RTA"] if "cash" in merged.index else 0.0)
    if cash_diff < 0:
        ws.cell(row=cash_row, column=2).font = red

    ws.append(["Total Diff", None])
    write_currency(ws, ws.max_row, 2, total_diff)

    # gray separator
    for row in range(1, ws.max_row + 1):
        ws[f"E{row}"].fill = gray_fill

    # Detailed recon (col G..)
    r = 1
    ws.cell(row=r, column=7).value = "Detailed Recon"; ws.cell(row=r, column=7).font = bold; r += 2
    ws.cell(row=r, column=7).value = "1881";           write_currency(ws, r, 8, round(batch_total, 2)); r += 1
    ws.cell(row=r, column=7).value = "RTA";            write_currency(ws, r, 8, rta_total);               r += 1
    ws.cell(row=r, column=7).value = "Diff";           write_currency(ws, r, 8, total_diff); write_currency(ws, r, 9, 0.00); r += 2
    ws.cell(row=r, column=7).value = "1881"; ws.cell(row=r, column=7).font = bold; r += 1
    for mode in batch_modes:
        ws.cell(row=r, column=7).value = mode.capitalize()
        write_currency(ws, r, 8, merged.loc[mode, "Amount_Bank"]); r += 1
    write_currency(ws, r, 8, round(batch_total, 2)); ws.cell(row=r, column=8).font = bold; r += 2

    # Unmatched tables
    ws.cell(row=r, column=7).value = "Transactions in 1881 not in RTA"; ws.cell(row=r, column=7).font = bold; r += 1
    headers_1 = ["Batch Date", "Card brand", "Card number", "Amount", "Matching", "Comments"]
    for col_idx, header in enumerate(headers_1, start=7):
        ws.cell(row=r, column=col_idx).value = header; ws.cell(row=r, column=col_idx).font = bold
    r += 1
    for _, row_data in batch_unmatched.iterrows():
        for col_idx, val in enumerate(row_data, start=7):
            ws.cell(row=r, column=col_idx).value = val
        write_currency(ws, r, 10, ws.cell(row=r, column=10).value); r += 1

    r += 1
    ws.cell(row=r, column=7).value = "Transactions in RTA not in 1881"; ws.cell(row=r, column=7).font = bold; r += 1
    # MODIFIED HEADERS_2 TO INCLUDE RegID
    headers_2 = ["Date/Time", "RegID", "Card brand", "Card number", "Amount", "Matching", "Comments"]
    for col_idx, header in enumerate(headers_2, start=7):
        ws.cell(row=r, column=col_idx).value = header; ws.cell(row=r, column=col_idx).font = bold
    r += 1
    for _, row_data in rta_unmatched.iterrows():
        for col_idx, val in enumerate(row_data, start=7):
            ws.cell(row=r, column=col_idx).value = val
        # CHANGED COLUMN INDEX FOR AMOUNT FROM 10 TO 11
        write_currency(ws, r, 11, ws.cell(row=r, column=11).value); r += 1

    # Difference section
    r += 2
    ws.cell(row=r, column=7).value = "Difference"; ws.cell(row=r, column=7).font = bold; r += 2
    headers_diff = ["Date", "Card / Tender", "Card number", "Amount"]
    for col_idx, header in enumerate(headers_diff, start=7):
        ws.cell(row=r, column=col_idx).value = header; ws.cell(row=r, column=col_idx).font = bold
    r += 1
    ws.cell(row=r, column=7).value = "1881"; ws.cell(row=r, column=7).font = bold; r += 1

    blank_batch_diff = batch_unmatched[(batch_unmatched["Matching"] == "") & (batch_unmatched["Comments"] == "")]
    for _, row_data in blank_batch_diff.iterrows():
        ws.cell(row=r, column=7).value = row_data["Batch Date"]
        brand_val = str(row_data["Card brand"]).capitalize() if pd.notna(row_data["Card brand"]) else ""
        ws.cell(row=r, column=8).value = brand_val
        ws.cell(row=r, column=9).value = row_data["Card number"]
        write_currency(ws, r, 10, row_data["Amount"]); r += 1

    ws.cell(row=r, column=7).value = "RTA"; ws.cell(row=r, column=7).font = bold; r += 1
    blank_rta_diff = rta_unmatched[(rta_unmatched["Matching"] == "") & (rta_unmatched["Comments"] == "")]
    for _, row_data in blank_rta_diff.iterrows():
        ws.cell(row=r, column=7).value = row_data["Date/Time"]
        brand_val = str(row_data["Card brand"]).capitalize() if pd.notna(row_data["Card brand"]) else ""
        ws.cell(row=r, column=8).value = brand_val
        ws.cell(row=r, column=9).value = row_data["Card number"]
        write_currency(ws, r, 10, row_data["Amount"]); r += 1

    # -----------------------------
    # Card Diff Initial + Category Diff + Diff Final (aligned rows)
    # -----------------------------
    r += 2
    card_header_row = r
    ws.cell(row=card_header_row, column=7).value = "Card"
    ws.cell(row=card_header_row, column=8).value = "Diff Initial"
    ws.cell(row=card_header_row, column=9).value = "Category Diff"
    ws.cell(row=card_header_row, column=10).value = "Diff Final"
    for col in range(7, 11):
        ws.cell(row=card_header_row, column=col).font = bold
        ws.cell(row=card_header_row, column=col).fill = blue_fill

    data_start = card_header_row + 1
    rowp = data_start

    for card in batch_modes:
        value = merged.loc[card, "Diff"]
        ws.cell(row=rowp, column=7).value = card.title()
        write_currency(ws, rowp, 8, value)
        rowp += 1

    # Cash row after batch_modes
    batch_cash = batch_df[batch_df["Card brand"].astype(str).str.lower() == "cash"]["Amount"].sum()
    rta_cash   = rta_df[rta_df["Tender"].astype(str).str.lower() == "cash"]["Total"].sum()
    cash_value = batch_cash - rta_cash
    ws.cell(row=rowp, column=7).value = "Cash"
    c = write_currency(ws, rowp, 8, cash_value)
    if cash_value < 0: c.font = red
    rowp += 1

    # Build card_brands for category-diff alignment (Title case)
    card_brands = [m.title() for m in batch_modes] + ["Cash"]
    category_diff_dict = {brand: 0 for brand in card_brands}

    def extract_brand_from_comment(comment):
        comment = str(comment).lower()
        if "matched with cash" in comment:
            return "Cash"
        for brand in card_brands:
            brand_lower = brand.lower()
            if f"categorize as {brand_lower}" in comment or brand_lower in comment:
                return brand
        return None

    # Sum adds/subtracts for Category Diff
    for _, rowx in batch_unmatched.iterrows():
        brand = extract_brand_from_comment(rowx.get("Comments", ""))
        if brand:
            category_diff_dict[brand] += rowx["Amount"]

    for _, rowx in rta_unmatched.iterrows():
        brand = extract_brand_from_comment(rowx.get("Comments", ""))
        if brand:
            category_diff_dict[brand] -= rowx["Amount"]

    # Write Category Diff and Diff Final aligned with Diff Initial rows
    diff_final_total = 0.0
    for i, brand in enumerate(card_brands):
        row_num = data_start + i
        di_val = ws.cell(row=row_num, column=8).value or 0.0
        cd_val = category_diff_dict.get(brand, 0.0)
        final_diff = float(di_val) + float(cd_val)
        c = write_currency(ws, row_num, 9, cd_val)
        if cd_val < 0: c.font = red
        c2 = write_currency(ws, row_num, 10, final_diff)
        if final_diff < 0: c2.font = red
        diff_final_total += final_diff

    # Totals row under the card_brands
    total_row = data_start + len(card_brands)
    ws.cell(row=total_row, column=9).value = "Total"; ws.cell(row=total_row, column=9).font = bold
    write_currency(ws, total_row, 10, diff_final_total); ws.cell(row=total_row, column=10).font = bold

    # Append logic-generated 0.00 beside Diff Final total like your original logic
    blank_batch_diff = batch_unmatched[(batch_unmatched["Matching"] == "") & (batch_unmatched["Comments"] == "")]
    blank_rta_diff   = rta_unmatched[(rta_unmatched["Matching"] == "") & (rta_unmatched["Comments"] == "")]
    batch_diff_sum = blank_batch_diff["Amount"].sum()
    rta_diff_sum = blank_rta_diff["Amount"].sum()
    logic_zero_result = diff_final_total - batch_diff_sum + rta_diff_sum
    write_currency(ws, total_row, 11, logic_zero_result); ws.cell(row=total_row, column=11).font = bold

    # -----------------------------
    # RTA section (col G/H) + logic rows (start after totals)
    # -----------------------------
    r = total_row + 2
    ws.cell(row=r, column=7).value = "RTA"; ws.cell(row=r, column=7).font = bold; r += 1

    card_map = {"american express": "AmEx", "mastercard": "MstrCard", "visa": "Visa", "discover": "Discover", 
                "split": "Split", "check": "Check", "account": "Account"}
    for mode in ["american express", "mastercard", "visa", "discover", "split", "check", "account"]:
        ws.cell(row=r, column=7).value = card_map.get(mode, mode.title())
        write_currency(ws, r, 8, merged.loc[mode, "Amount_RTA"] if mode in merged.index else 0.0)
        r += 1

    # Cash row
    ws.cell(row=r, column=7).value = "Cash"; write_currency(ws, r, 8, merged.loc["cash", "Amount_RTA"] if "cash" in merged.index else 0.0); r += 1

    # Totals + logic rows (same formulae as before)
    ws.cell(row=r, column=7).value = "Total"
    c = write_currency(ws, r, 8, rta_total)
    ws.cell(row=r, column=7).font = bold; ws.cell(row=r, column=8).font = bold

    unmatched_batch_total = blank_batch_diff["Amount"].sum()
    unmatched_rta_total   = blank_rta_diff["Amount"].sum()
    final_logic_value     = batch_total - unmatched_batch_total + unmatched_rta_total
    write_currency(ws, r, 9, final_logic_value); ws.cell(row=r, column=9).font = bold
    diff_total = round(rta_total - final_logic_value, 2)
    write_currency(ws, r, 10, diff_total); ws.cell(row=r, column=10).font = bold
    r += 1
    autofit_columns(ws)

    # =========================
    # NEW TAB: Matched Transactions Section
    # =========================
    ws2 = wb.create_sheet("Matched Transactions")
    headers_tab2 = ["Batch Date", "Card brand", "Card number", "Amount", "RTA RegID", "RTA Date/Time"]
    for col_idx, header in enumerate(headers_tab2, start=1):
        ws2.cell(row=1, column=col_idx).value = header
        ws2.cell(row=1, column=col_idx).font = bold
        ws2.cell(row=1, column=col_idx).fill = blue_fill
    
    r2 = 2
    for item in exact_matches_for_tab2:
        ws2.cell(row=r2, column=1).value = item["Batch Date"]
        ws2.cell(row=r2, column=2).value = item["Card brand"]
        ws2.cell(row=r2, column=3).value = item["Card number"]
        write_currency(ws2, r2, 4, item["Amount"])
        ws2.cell(row=r2, column=5).value = item["RTA RegID"]
        ws2.cell(row=r2, column=6).value = item["RTA Date/Time"]
        r2 += 1
    
    autofit_columns(ws2)

    if output_file is None:
        output_file = f"Final_Bank_Recon_Combined_{target_date}.xlsx"
    wb.save(output_file)

    return output_file

# Example runner (optional) ---------------------------------------------------
if __name__ == "__main__":
    # replace these with your actual paths or call run_reconciliation from another script
    batch_path = r"C:\Users\ADMIN\Documents\ONHO\Batch Aug.xlsx"
    rta_path = r"C:\Users\ADMIN\Documents\ONHO\RTA.xlsx"
    out = run_reconciliation(batch_path, rta_path)
    print("Saved:", out)
